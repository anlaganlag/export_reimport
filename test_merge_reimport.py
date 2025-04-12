#!/usr/bin/env python3
import os
import sys
import subprocess
from openpyxl import Workbook, load_workbook

def create_test_files():
    """Create test Excel files for testing merge_reimport.py"""
    # Create output directory if it doesn't exist
    if not os.path.exists('test_output'):
        os.makedirs('test_output')
    
    # Create header files
    pl_header = Workbook()
    pl_header_sheet = pl_header.active
    pl_header_sheet['A1'] = 'Packing List Header'
    pl_header_sheet['B1'] = 'Header Info'
    pl_header_sheet.merge_cells('A1:B1')
    pl_header.save('test_output/pl_header.xlsx')
    
    ci_header = Workbook()
    ci_header_sheet = ci_header.active
    ci_header_sheet['A1'] = 'Commercial Invoice Header'
    ci_header_sheet['B1'] = 'Header Info'
    ci_header_sheet.merge_cells('A1:B1')
    ci_header.save('test_output/ci_header.xlsx')
    
    # Create footer files
    pl_footer = Workbook()
    pl_footer_sheet = pl_footer.active
    pl_footer_sheet['A1'] = 'Packing List Footer'
    pl_footer_sheet['B1'] = 'Footer Info'
    pl_footer_sheet.merge_cells('A1:B1')
    pl_footer.save('test_output/pl_footer.xlsx')
    
    ci_footer = Workbook()
    ci_footer_sheet = ci_footer.active
    ci_footer_sheet['A1'] = 'Commercial Invoice Footer'
    ci_footer_sheet['B1'] = 'Footer Info'
    ci_footer_sheet.merge_cells('A1:B1')
    ci_footer.save('test_output/ci_footer.xlsx')
    
    # Create main file with multiple sheets
    main_file = Workbook()
    
    # Create Packing List sheet
    pl_sheet = main_file.active
    pl_sheet.title = 'Packing List'
    pl_sheet['A1'] = 'Packing List Content'
    pl_sheet['B1'] = 'Item 1'
    pl_sheet['A2'] = 'Item 2'
    pl_sheet['B2'] = 'Details'
    pl_sheet.merge_cells('A2:B2')
    
    # Create Commercial Invoice sheet
    ci_sheet = main_file.create_sheet('Commercial Invoice')
    ci_sheet['A1'] = 'Commercial Invoice Content'
    ci_sheet['B1'] = 'Item 1'
    ci_sheet['A2'] = 'Item 2'
    ci_sheet['B2'] = 'Details'
    ci_sheet.merge_cells('A2:B2')
    
    main_file.save('test_output/main_file.xlsx')
    
    return {
        'pl_header': os.path.abspath('test_output/pl_header.xlsx'),
        'pl_footer': os.path.abspath('test_output/pl_footer.xlsx'),
        'ci_header': os.path.abspath('test_output/ci_header.xlsx'),
        'ci_footer': os.path.abspath('test_output/ci_footer.xlsx'),
        'main_file': os.path.abspath('test_output/main_file.xlsx'),
        'output_file': os.path.abspath('test_output/output_file.xlsx')
    }

def test_merge_reimport():
    """Test the merge_reimport.py script"""
    # Create test files
    test_files = create_test_files()
    
    # Get the path to merge_reimport.py
    script_dir = os.path.dirname(os.path.abspath(__file__))
    merge_reimport_path = os.path.join(script_dir, 'merge_reimport.py')
    
    if not os.path.exists(merge_reimport_path):
        print(f"Error: merge_reimport.py not found at {merge_reimport_path}")
        return False
    
    # Run merge_reimport.py
    cmd = [
        sys.executable,
        merge_reimport_path,
        test_files['main_file'],
        test_files['output_file'],
        test_files['pl_header'],
        test_files['pl_footer'],
        test_files['ci_header'],
        test_files['ci_footer']
    ]
    
    print(f"Running command: {' '.join(cmd)}")
    
    try:
        # Use shell=True to avoid issues with spaces in paths
        cmd_str = ' '.join([f'"{arg}"' if ' ' in str(arg) else str(arg) for arg in cmd])
        result = subprocess.run(cmd_str, shell=True, check=True, capture_output=True, text=True)
        print("Command output:")
        print(result.stdout)
        
        if result.stderr:
            print("Command errors:")
            print(result.stderr)
        
        # Check if the output file was created
        if not os.path.exists(test_files['output_file']):
            print("Error: Output file was not created")
            return False
        
        # Check the content of the output file
        output_wb = load_workbook(test_files['output_file'])
        
        # Check Packing List sheet
        if 'Packing List' not in output_wb.sheetnames:
            print("Error: Packing List sheet not found in output file")
            return False
        
        pl_sheet = output_wb['Packing List']
        
        # Check if header, content, and footer were merged correctly
        if pl_sheet['A1'].value != 'Packing List Header':
            print("Error: Header not merged correctly in Packing List sheet")
            return False
        
        if pl_sheet['A2'].value != 'Packing List Content':
            print("Error: Content not merged correctly in Packing List sheet")
            return False
        
        if pl_sheet['A3'].value != 'Packing List Footer':
            print("Error: Footer not merged correctly in Packing List sheet")
            return False
        
        # Check Commercial Invoice sheet
        if 'Commercial Invoice' not in output_wb.sheetnames:
            print("Error: Commercial Invoice sheet not found in output file")
            return False
        
        ci_sheet = output_wb['Commercial Invoice']
        
        # Check if header, content, and footer were merged correctly
        if ci_sheet['A1'].value != 'Commercial Invoice Header':
            print("Error: Header not merged correctly in Commercial Invoice sheet")
            return False
        
        if ci_sheet['A2'].value != 'Commercial Invoice Content':
            print("Error: Content not merged correctly in Commercial Invoice sheet")
            return False
        
        if ci_sheet['A3'].value != 'Commercial Invoice Footer':
            print("Error: Footer not merged correctly in Commercial Invoice sheet")
            return False
        
        print("Test passed successfully!")
        return True
    
    except subprocess.CalledProcessError as e:
        print(f"Error running merge_reimport.py: {e}")
        if e.stdout:
            print("Command output:")
            print(e.stdout)
        if e.stderr:
            print("Command errors:")
            print(e.stderr)
        return False
    except Exception as e:
        print(f"Error: {e}")
        return False

if __name__ == "__main__":
    test_merge_reimport() 