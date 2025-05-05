#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Module for Excel styling and formatting functions.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
import re

def apply_font_style(cell, is_bold=False):
    """
    Apply consistent font styling to a cell.
    This is extracted from the original process_shipping_list.py.

    Args:
        cell: The openpyxl cell to style
        is_bold (bool): Whether to make the text bold
    """
    cell.font = Font(name='Arial', size=10, bold=is_bold)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def apply_selective_bold(ws):
    """
    Apply bold formatting to specific cells based on content.
    This is extracted from the original process_shipping_list.py.

    Args:
        ws: Worksheet to apply formatting to
    """
    # Bold for title row and key cells
    bold_patterns = [
        'PACKING LIST', '装箱单', 'COMMERCIAL INVOICE', '商业发票',
        'INVOICE', 'PROFORMA', 'REIMPORT', 'IMPORT',
        'CIFC SHENZHEN', 'FOB', 'CIF',
        'TOTAL', '总计', '合计',
        '^QUANTITY$', '^数量$', '^QTY$', '^NET$', '^GROSS$',
        'DESCRIPTION', '描述', '产品描述', 'PRODUCT', '产品',
    ]
    
    # Regular expression patterns compiled for efficiency
    bold_regexes = [re.compile(pattern, re.IGNORECASE) for pattern in bold_patterns]
    
    # Scan all cells and apply bold where pattern matches
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Check if cell value matches any of our patterns
                if any(regex.search(cell.value) for regex in bold_regexes):
                    cell.font = Font(name='Arial', size=10, bold=True)
                    
                    # For titles, also center and make larger
                    if any(p in cell.value.upper() for p in ['PACKING LIST', 'COMMERCIAL INVOICE', 'INVOICE', '装箱单', '商业发票']):
                        cell.font = Font(name='Arial', size=14, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

def apply_pl_footer_styling(workbook_path):
    """
    Apply styling to Packing List footer.
    This is extracted from the original process_shipping_list.py.

    Args:
        workbook_path: Path to the Excel workbook to style
    """
    # Load the workbook
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active  # Assuming the first sheet is the packing list
    
    # Define border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Find the footer section (typically contains "TOTAL" or similar)
    footer_row = None
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and ('TOTAL' in cell.value.upper() or '合计' in cell.value):
                footer_row = row_idx
                break
        if footer_row:
            break
            
    if footer_row:
        # Apply styles to the footer row
        for cell in ws[footer_row]:
            cell.border = thin_border
            cell.font = Font(name='Arial', size=10, bold=True)
            
            # Apply gray background to total cells
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
    # Save the styled workbook
    wb.save(workbook_path)

def apply_excel_styling(file_path):
    """
    Apply comprehensive styling to an Excel file.
    This is extracted from the original process_shipping_list.py.

    Args:
        file_path: Path to the Excel file to style
    """
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Style each worksheet
    for ws in wb.worksheets:
        # Set column widths based on content
        column_widths = {
            'Material code': 35,
            'Unit Price': 20,
            'DESCRIPTION': 30,
            'Part': 25,
            'Part No': 25,
            'No': 10,
            'Quantity': 15,
            'QTY': 15,
            'Net': 15,
            'Gross': 15,
            'Unit': 10,
            'Amount': 20,
            'FOB': 20,
            'CIF': 20,
        }
        
        # Apply column widths
        for column_cells in ws.columns:
            col = column_cells[0].column_letter  # Get column letter
            header = ws[f"{col}1"].value
            
            if header in column_widths:
                ws.column_dimensions[col].width = column_widths[header]
            elif header and any(key in str(header) for key in column_widths):
                for key, width in column_widths.items():
                    if key in str(header):
                        ws.column_dimensions[col].width = width
                        break
            else:
                # Default width if no match
                ws.column_dimensions[col].width = 15
        
        # Apply font styling to all cells
        for row in ws.iter_rows():
            for cell in row:
                apply_font_style(cell)
        
        # Apply selective bold formatting
        apply_selective_bold(ws)
        
        # Set uniform row height
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
        
        # Apply border to all data cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Highlight header row
        for cell in ws[1]:
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Save the styled workbook
    wb.save(file_path) 