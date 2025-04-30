# -*- coding: utf-8 -*-
import openpyxl
import copy
import os
import sys

def copy_cell_formatting(source_cell, target_cell):
    """Helper function to copy cell formatting."""
    if hasattr(source_cell, '_style'):
        target_cell._style = copy.copy(source_cell._style)
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)

def copy_sheet(source_sheet, target_sheet):
    """Copy contents and formatting from source sheet to target sheet."""
    # Copy cell values and formatting
    for row_idx, row in enumerate(source_sheet.rows, 1):
        for col_idx, source_cell in enumerate(row, 1):
            target_cell = target_sheet.cell(row=row_idx, column=col_idx)

            if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                target_cell.value = source_cell.value
                copy_cell_formatting(source_cell, target_cell)

    # Copy merged cells
    for merged_range in source_sheet.merged_cells:
        merge_range = f"{openpyxl.utils.get_column_letter(merged_range.min_col)}{merged_range.min_row}:{openpyxl.utils.get_column_letter(merged_range.max_col)}{merged_range.max_row}"
        try:
            target_sheet.merge_cells(merge_range)
        except ValueError as e:
            print(f"Warning: Could not merge cells {merge_range}: {e}")

    # Copy column dimensions
    for col_idx, column in enumerate(source_sheet.columns, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

    # Copy row dimensions
    for row_idx in range(1, source_sheet.max_row + 1):
        if row_idx in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row_idx].height = source_sheet.row_dimensions[row_idx].height

def append_sheet_with_offset(source_sheet, target_sheet, row_offset, filename):
    """Append source sheet to target sheet with row offset."""
    # Copy data with offset
    for row_idx, row in enumerate(source_sheet.rows, 1):
        for col_idx, source_cell in enumerate(row, 1):
            target_cell = target_sheet.cell(row=row_offset + row_idx, column=col_idx)

            if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                target_cell.value = source_cell.value
                copy_cell_formatting(source_cell, target_cell)

    # Process merged cells with offset
    for merged_range in source_sheet.merged_cells:
        min_row = merged_range.min_row + row_offset
        max_row = merged_range.max_row + row_offset
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        merge_range = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

        try:
            target_sheet.merge_cells(merge_range)
            source_value = source_sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            target_sheet.cell(row=min_row, column=min_col).value = source_value
        except ValueError as e:
            print(f"Warning: Could not merge cells {merge_range}: {e}")

    return source_sheet.max_row

def apply_column_widths(sheet, column_widths, default_widths=None):
    """Apply column widths to sheet based on header names."""
    default_widths = default_widths or {
        'Material code': 35,
        'Unit Price': 20,
        'DESCRIPTION': 30,
        'default': 15
    }

    for col_idx, cell in enumerate(sheet[1], 1):
        col_name = cell.value
        col_letter = openpyxl.utils.get_column_letter(col_idx)

        if col_name in column_widths:
            sheet.column_dimensions[col_letter].width = column_widths[col_name]
        elif col_name in default_widths:
            sheet.column_dimensions[col_letter].width = default_widths[col_name]
        else:
            sheet.column_dimensions[col_letter].width = default_widths['default']

def load_workbook_safely(file_path):
    """Load workbook with error handling."""
    try:
        return openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error opening file {file_path}: {e}")
        return None

def merge_three_excel_files(first_file, middle_file, last_file, output_file, first_sheet_first_file=None, first_sheet_last_file=None):
    """
    Merge Excel files with specific requirements:
    - First file used for headers (h.xlsx)
    - Middle file - first sheet is PL, remaining sheets are invoices
    - Last file used for footers (f.xlsx)
    - All merged cells and formatting preserved

    If first_sheet_first_file and first_sheet_last_file are provided, they will be used
    to merge with the first sheet (Packing List) of the middle file.
    """
    print(f"Merging files: {first_file}, {middle_file}, {last_file}")

    # Process middle file first to extract all sheets
    middle_wb = load_workbook_safely(middle_file)
    if not middle_wb or len(middle_wb.sheetnames) < 1:
        print(f"Error: Middle file must have at least 1 sheet")
        return False

    # Get original sheet names from middle file to preserve them
    pl_sheet_name = middle_wb.sheetnames[0]  # First sheet (Packing List)
    invoice_sheet_names = middle_wb.sheetnames[1:]  # All remaining sheets (Invoices)

    print(f"Found sheets in middle file:")
    print(f"- Packing List: '{pl_sheet_name}'")
    print(f"- Invoice sheets: {invoice_sheet_names}")

    # Create output workbook
    merged_wb = openpyxl.Workbook()
    
    # Create and copy Packing List from middle file's first sheet
    packing_list_sheet = merged_wb.active
    packing_list_sheet.title = pl_sheet_name

    # Extract column widths from middle file's first sheet
    middle_pl_sheet = middle_wb[pl_sheet_name]
    pl_column_widths = {}
    for col_idx, cell in enumerate(middle_pl_sheet[1], 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if cell.value and col_letter in middle_pl_sheet.column_dimensions:
            pl_column_widths[cell.value] = middle_pl_sheet.column_dimensions[col_letter].width

    # Handle Packing List sheet merging if first_sheet_first_file and first_sheet_last_file are provided
    if first_sheet_first_file and first_sheet_last_file:
        print(f"Merging {pl_sheet_name} with: {first_sheet_first_file}, {middle_file}, {first_sheet_last_file}")

        # Prepare data for merging Packing List
        pl_files_to_merge = [
            (first_sheet_first_file, lambda wb: wb.active),
            (middle_file, lambda wb: wb[wb.sheetnames[0]]),
            (first_sheet_last_file, lambda wb: wb.active)
        ]

        # Merge Packing List sheets vertically
        pl_row_offset = 0
        for file_path, sheet_selector in pl_files_to_merge:
            wb = load_workbook_safely(file_path)
            if not wb:
                return False

            sheet = sheet_selector(wb)
            print(f"Processing {pl_sheet_name}: {file_path}")
            pl_row_offset += append_sheet_with_offset(sheet, packing_list_sheet, pl_row_offset, file_path)

        # Apply column widths to Packing List
        apply_column_widths(packing_list_sheet, pl_column_widths)
    else:
        # If no first sheet files provided, just copy the first sheet from middle file
        print(f"Copying first sheet from {middle_file}")
        copy_sheet(middle_wb[pl_sheet_name], packing_list_sheet)

    # Process each invoice sheet
    for invoice_sheet_name in invoice_sheet_names:
        print(f"\nProcessing invoice sheet: {invoice_sheet_name}")
        
        # Create new sheet for this invoice
        invoice_sheet = merged_wb.create_sheet(invoice_sheet_name)
        
        # Extract column widths from middle file's invoice sheet
        middle_invoice_sheet = middle_wb[invoice_sheet_name]
        invoice_column_widths = {}
        for col_idx, cell in enumerate(middle_invoice_sheet[1], 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if cell.value and col_letter in middle_invoice_sheet.column_dimensions:
                invoice_column_widths[cell.value] = middle_invoice_sheet.column_dimensions[col_letter].width

        # Load first file (h.xlsx) and get its second sheet for header
        first_wb = load_workbook_safely(first_file)
        if not first_wb or len(first_wb.sheetnames) < 2:
            print(f"Error: First file (h.xlsx) must have at least 2 sheets")
            return False

        # Get the second sheet from h.xlsx
        h_invoice_sheet = first_wb[first_wb.sheetnames[1]]
        
        # Load last file (f.xlsx) for footer
        last_wb = load_workbook_safely(last_file)
        if not last_wb:
            print(f"Error: Could not load last file (f.xlsx)")
            return False
            
        f_invoice_sheet = last_wb.active

        # Merge sheets vertically for this invoice
        row_offset = 0
        
        # Add header from h.xlsx (second sheet)
        print(f"Adding header from {first_file} (second sheet)")
        row_offset += append_sheet_with_offset(h_invoice_sheet, invoice_sheet, row_offset, first_file)
        
        # Add content from middle file
        print(f"Adding content from {middle_file} ({invoice_sheet_name})")
        row_offset += append_sheet_with_offset(middle_invoice_sheet, invoice_sheet, row_offset, middle_file)
        
        # Add footer from f.xlsx
        print(f"Adding footer from {last_file}")
        row_offset += append_sheet_with_offset(f_invoice_sheet, invoice_sheet, row_offset, last_file)
        
        # Apply column widths to this invoice sheet
        apply_column_widths(invoice_sheet, invoice_column_widths)

    # Save result
    try:
        # Reorder sheets - PL first, then all invoice sheets
        sheet_order = [pl_sheet_name] + invoice_sheet_names
        merged_wb._sheets = [merged_wb[name] for name in sheet_order]
        
        merged_wb.save(output_file)
        print(f"\nSuccessfully saved merged file to: {output_file}")
        print(f"Final sheet order: {sheet_order}")
        return True
    except Exception as e:
        print(f"Error saving output file {output_file}: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Usage: python merge.py <first_file.xlsx> <middle_file.xlsx> <last_file.xlsx> <output_file.xlsx> [first_sheet_first_file.xlsx] [first_sheet_last_file.xlsx]")
        # Return error code but don't exit the process
        sys.exit(1)

    try:
        files = [os.path.abspath(sys.argv[i]) for i in range(1, 5)]

        # Check if first sheet files are provided
        first_sheet_first_file = os.path.abspath(sys.argv[5]) if len(sys.argv) > 5 else None
        first_sheet_last_file = os.path.abspath(sys.argv[6]) if len(sys.argv) > 6 else None

        success = merge_three_excel_files(
            files[0], files[1], files[2], files[3],
            first_sheet_first_file, first_sheet_last_file
        )

        if not success:
            print("Merge operation failed!")
            # Return error code
            sys.exit(1)
    except Exception as e:
        print(f"Error during merge operation: {e}")
        # Return error code
        sys.exit(1)

# 在Windows系统下自动打开合并后的Excel文件
# if os.name == 'nt':
#     os.startfile(output_file)
#     print("Opening merged Excel file...")
#     print("按回车键退出程序...")
#     input()