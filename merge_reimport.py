#!/usr/bin/env python3
import sys
import os
import openpyxl
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def copy_cell_formatting(source_cell, target_cell):
    """Helper function to copy cell formatting."""
    if hasattr(source_cell, '_style'):
        target_cell._style = copy.copy(source_cell._style)
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)

def copy_sheet(source_sheet, target_sheet):
    """Copy contents and formatting from source sheet to target sheet."""
    # Copy cell values and formatting
    for row_idx, row in enumerate(source_sheet.rows, 1):
        for col_idx, source_cell in enumerate(row, 1):
            target_cell = target_sheet.cell(row=row_idx, column=col_idx)
            
            if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                target_cell.value = source_cell.value
                copy_cell_formatting(source_cell, target_cell)
    
    # Copy merged cells
    for merged_range in source_sheet.merged_cells:
        merge_range = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}:{get_column_letter(merged_range.max_col)}{merged_range.max_row}"
        try:
            target_sheet.merge_cells(merge_range)
        except ValueError as e:
            print(f"Warning: Could not merge cells {merge_range}: {e}")
    
    # Copy column dimensions
    for col_idx, column in enumerate(source_sheet.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width
    
    # Copy row dimensions
    for row_idx in range(1, source_sheet.max_row + 1):
        if row_idx in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row_idx].height = source_sheet.row_dimensions[row_idx].height

def append_sheet_with_offset(source_sheet, target_sheet, row_offset, filename):
    """Append source sheet to target sheet with row offset."""
    # Copy data with offset
    for row_idx, row in enumerate(source_sheet.rows, 1):
        for col_idx, source_cell in enumerate(row, 1):
            target_cell = target_sheet.cell(row=row_offset + row_idx, column=col_idx)
            
            if not isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                target_cell.value = source_cell.value
                copy_cell_formatting(source_cell, target_cell)
    
    # Process merged cells with offset
    for merged_range in source_sheet.merged_cells:
        min_row = merged_range.min_row + row_offset
        max_row = merged_range.max_row + row_offset
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        
        merge_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        
        try:
            target_sheet.merge_cells(merge_range)
            source_value = source_sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            target_sheet.cell(row=min_row, column=min_col).value = source_value
        except ValueError as e:
            print(f"Warning: Could not merge cells {merge_range}: {e}")
    
    return source_sheet.max_row

def apply_column_widths(sheet, column_widths, default_widths=None):
    """Apply column widths to sheet based on header names."""
    default_widths = default_widths or {
        'Material code': 35,
        'Unit Price': 20,
        'DESCRIPTION': 30,
        'Model NO.': 20,
        'default': 15
    }
    
    for col_idx, cell in enumerate(sheet[1], 1):
        col_name = cell.value
        col_letter = get_column_letter(col_idx)
        
        if col_name in column_widths:
            sheet.column_dimensions[col_letter].width = column_widths[col_name]
        elif col_name in default_widths:
            sheet.column_dimensions[col_letter].width = default_widths[col_name]
        else:
            sheet.column_dimensions[col_letter].width = default_widths['default']

def load_workbook_safely(file_path):
    """Load workbook with error handling."""
    try:
        return load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error opening file {file_path}: {e}")
        return None

def merge_sheets(header_file, main_file, footer_file, output_file, sheet_name, output_wb=None):
    """Merge header, main content, and footer into a single sheet."""
    try:
        # Load workbooks safely
        header_wb = load_workbook_safely(header_file)
        main_wb = load_workbook_safely(main_file)
        footer_wb = load_workbook_safely(footer_file)
        
        if not all([header_wb, main_wb, footer_wb]):
            return None
        
        # Get the sheets
        header_sheet = header_wb.active
        main_sheet = main_wb[sheet_name]
        footer_sheet = footer_wb.active
        
        # Create or get output workbook and sheet
        if output_wb is None:
            output_wb = Workbook()
            output_sheet = output_wb.active
            output_sheet.title = sheet_name
        else:
            # Create a new sheet in the existing workbook
            output_sheet = output_wb.create_sheet(sheet_name)
        
        # Extract column widths from main sheet
        column_widths = {}
        for col_idx, cell in enumerate(main_sheet[1], 1):
            col_letter = get_column_letter(col_idx)
            if cell.value and col_letter in main_sheet.column_dimensions:
                column_widths[cell.value] = main_sheet.column_dimensions[col_letter].width
        
        # Merge sheets vertically with offset
        row_offset = 0
        for source_sheet in [header_sheet, main_sheet, footer_sheet]:
            row_offset += append_sheet_with_offset(source_sheet, output_sheet, row_offset, '')
        
        # Apply column widths
        apply_column_widths(output_sheet, column_widths)
        
        return output_wb
    except Exception as e:
        print(f"Error merging sheets: {e}")
        return None

def main():
    if len(sys.argv) != 7:
        print("Usage: python merge_reimport.py <input_file> <output_file> <pl_header_file> <pl_footer_file> <ci_header_file> <ci_footer_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    pl_header_file = sys.argv[3]
    pl_footer_file = sys.argv[4]
    ci_header_file = sys.argv[5]
    ci_footer_file = sys.argv[6]
    
    print(f"Processing with following files:")
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print(f"PL Header: {pl_header_file}")
    print(f"PL Footer: {pl_footer_file}")
    print(f"CI Header: {ci_header_file}")
    print(f"CI Footer: {ci_footer_file}")
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Error: Input file {input_file} does not exist")
        sys.exit(1)
    
    # Load the input workbook to get sheet names
    print(f"Loading input workbook: {input_file}")
    input_wb = load_workbook_safely(input_file)
    if not input_wb:
        print(f"Failed to load input workbook: {input_file}")
        sys.exit(1)
    
    sheet_names = input_wb.sheetnames
    print(f"Found sheets: {sheet_names}")
    
    # Create output workbook
    output_wb = None
    
    # Process each sheet
    for sheet_name in sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        
        # Determine which header and footer files to use
        if sheet_name == 'Packing List':
            header_file = pl_header_file
            footer_file = pl_footer_file
            print(f"Using PL header and footer files")
        else:
            header_file = ci_header_file
            footer_file = ci_footer_file
            print(f"Using CI header and footer files")
        
        # Skip if header or footer files are missing or empty
        if not header_file or not footer_file:
            print(f"Warning: Header or footer files not provided for sheet {sheet_name}, skipping")
            continue
            
        if not os.path.exists(header_file) or not os.path.exists(footer_file):
            print(f"Warning: Header or footer files missing for sheet {sheet_name}, skipping")
            continue
        
        # Merge the sheets
        try:
            print(f"Merging sheets for {sheet_name}...")
            output_wb = merge_sheets(header_file, input_file, footer_file, output_file, sheet_name, output_wb)
            if output_wb:
                print(f"Successfully merged sheet: {sheet_name}")
            else:
                print(f"Failed to merge sheet: {sheet_name}")
                sys.exit(1)
        except Exception as e:
            print(f"Error merging sheet {sheet_name}: {str(e)}")
            import traceback
            print("Full error:")
            print(traceback.format_exc())
            sys.exit(1)
    
    # Reorder sheets to match expected order
    if output_wb and 'Packing List' in output_wb.sheetnames and 'Commercial Invoice' in output_wb.sheetnames:
        print("Reordering sheets...")
        output_wb._sheets = [output_wb['Packing List'], output_wb['Commercial Invoice']]
    
    # Save the output workbook
    try:
        print(f"Saving output workbook to: {output_file}")
        output_wb.save(output_file)
        print(f"Successfully processed all sheets in {input_file}")
    except Exception as e:
        print(f"Error saving output file {output_file}: {str(e)}")
        import traceback
        print("Full error:")
        print(traceback.format_exc())
        sys.exit(1)

if __name__ == "__main__":
    main()