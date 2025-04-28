import pandas as pd
import os
import time
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import shutil
import logging
import datetime

# Make sure outputs directory exists
if not os.path.exists('outputs'):
    try:
        os.makedirs('outputs')
        print("Created outputs directory")
    except Exception as e:
        print(f"Error creating outputs directory: {e}")
        raise

def apply_font_style(cell, is_bold=False):
    """Helper function to apply font style to a cell."""
    current_font = cell.font
    return Font(
        name=current_font.name if current_font.name else 'Arial',
        size=current_font.size if current_font.size else 11,
        bold=is_bold,
        italic=current_font.italic if current_font.italic else False,
        color=current_font.color if current_font.color else None
    )

def apply_selective_bold(ws):
    """Apply selective bold formatting to the worksheet."""
    # Make header row bold
    for cell in ws[1]:
        if cell.value:
            cell.font = apply_font_style(cell, is_bold=True)
    
    # Make all data rows not bold by default
    for row_idx in range(2, ws.max_row + 1):
        for cell in ws[row_idx]:
            if cell.value:
                cell.font = apply_font_style(cell, is_bold=False)
    
    # Find the Total row
    total_row = None
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=2).value == "Total":
            total_row = row_idx
            break
    
    if total_row:
        # Make Total row and subsequent rows bold
        for row_idx in range(total_row, ws.max_row + 1):
            for cell in ws[row_idx]:
                if cell.value:
                    cell.font = apply_font_style(cell, is_bold=True)

# Function to read Excel files
def read_excel_file(file_path, skip=0):
    """
    读取Excel文件，处理多层表头
    
    Args:
        file_path: Excel文件路径
        skip: 要跳过的行数(用于跳过表头)
    
    Returns:
        pd.DataFrame: 加载的数据
    """
    try:
        # Try reading with multi-level headers (English + Chinese)
        # First row (0) is table title, second row (1) is English headers, 
        # third row (2) is Chinese headers, data starts at row 4
        print(f"Reading Excel file: {file_path}")
        
        # 修正：原装箱单结构 - 第一行是标题，第二行是英文表头，第三行是中文表头
        # 如果要读取数据行，应从第四行开始(索引为3)，所以skip应为2(跳过前两行)而不是3
        
        # If skip is provided, use it (used to skip specific number of rows)
        if skip > 0:
            # 修正skip=3的情况 - 应该改为skip=2以确保从正确的数据行开始
            if skip == 3:
                print(f"Converting skip=3 to skip=2 to prevent skipping first data row")
                skip = 2
                
            print(f"Skipping {skip} rows as specified")
            return pd.read_excel(file_path, skiprows=skip)
            
        # Default behavior for packing lists - handle multi-level headers
        # Read first few rows to check structure
        header_peek = pd.read_excel(file_path, nrows=4)
        print(f"First 4 rows preview:")
        for i, row in enumerate(header_peek.values.tolist()):
            print(f"  Row {i+1}: {row[:5]}...")
        
        # Use multi-level headers (English row + Chinese row)
        # Skip the first row (table title)
        df = pd.read_excel(file_path, header=[1, 2], skiprows=[0])
        
        # Debug column names
        print(f"Column names after reading with multi-level headers: {df.columns.tolist()[:5]}...")
        
        # Convert multi-level columns to single level for easier processing
        # Combine English and Chinese header names with a separator
        df.columns = [f"{col[0]}|{col[1]}" if isinstance(col, tuple) and len(col) > 1 
                      else col for col in df.columns]
        
        print(f"Simplified column names: {df.columns.tolist()[:5]}...")
        return df
        
    except Exception as e:
        print(f"Error reading with multi-level headers: {e}")
        print("Falling back to standard Excel reading...")
        # Fallback to standard reading
        return pd.read_excel(file_path)




def merge_packing_list_cells(workbook_path):
    """
    Merge cells in the Packing List sheet for rows with the same Carton NO.
    Specifically merges the CTNS, Carton MEASUREMENT, G.W (KG), and Carton NO. columns vertically,
    but only for groups with more than one row.
    """
    try:
        wb = load_workbook(workbook_path)
        if 'Packing List' not in wb.sheetnames and 'PL' not in wb.sheetnames:
            print("No 'Packing List' or 'PL' sheet found in workbook")
            return False
            
        # 使用'PL'或'Packing List'作为工作表名称
        sheet_name = 'PL' if 'PL' in wb.sheetnames else 'Packing List'
        ws = wb[sheet_name]
        
        # Find column indices
        carton_no_idx = None
        ctns_idx = None
        measurement_idx = None
        gw_idx = None
        desc_idx = None
        
        # 新旧列名映射关系
        column_name_mapping = {
            'Carton Number': ['Carton NO.', 'Carton Number', '箱号'],
            'Total Carton Quantity': ['CTNS', 'Total Carton Quantity', '件数'],
            'Total Volume (CBM)': ['Carton MEASUREMENT', 'Total Volume (CBM)', '体积'],
            'Total Gross Weight (kg)': ['G.W (KG)', 'Total Gross Weight (kg)', '毛重'],
            '名称': ['DESCRIPTION', '名称', '描述']
        }
        
        for col_idx, cell in enumerate(ws[1], 1):
            cell_value = cell.value
            if not cell_value:
                continue
                
            cell_value_str = str(cell_value).strip()
            
            # 使用映射关系查找对应的列
            for target_col, possible_names in column_name_mapping.items():
                if cell_value_str in possible_names or any(name.lower() in cell_value_str.lower() for name in possible_names):
                    if target_col == 'Carton Number':
                        carton_no_idx = col_idx
                    elif target_col == 'Total Carton Quantity':
                        ctns_idx = col_idx
                    elif target_col == 'Total Volume (CBM)':
                        measurement_idx = col_idx
                    elif target_col == 'Total Gross Weight (kg)':
                        gw_idx = col_idx
                    elif target_col == '名称':
                        desc_idx = col_idx
        
        if not all([carton_no_idx, ctns_idx, measurement_idx, gw_idx]):
            print("Could not find all required columns for merging")
            print(f"Found: Carton NO: {carton_no_idx}, CTNS: {ctns_idx}, Measurement: {measurement_idx}, G.W: {gw_idx}")
            return False
            
        # Track rows with the same carton number
        current_carton = None
        start_row = None
        last_row = ws.max_row
        
        # Check if the last row is a "Total" row
        total_row_idx = None
        for row_idx in range(last_row, 1, -1):
            if desc_idx and ws.cell(row=row_idx, column=desc_idx).value in ['Total', 'total', '合计', '总计']:
                total_row_idx = row_idx
                break
        
        # If we found a Total row, adjust the last_row to be one before it
        effective_last_row = total_row_idx - 1 if total_row_idx else last_row
        
        for row_idx in range(2, effective_last_row + 1):
            carton_no = ws.cell(row=row_idx, column=carton_no_idx).value
            
            # Skip empty or None values
            if not carton_no:
                continue
            
            if carton_no != current_carton:
                # 处理上一组（如果有）
                if start_row and start_row < row_idx - 1:  # 如果上一组有多行
                    end_row = row_idx - 1
                    ws.merge_cells(start_row=start_row, start_column=ctns_idx, 
                                  end_row=end_row, end_column=ctns_idx)
                    ws.merge_cells(start_row=start_row, start_column=measurement_idx,
                                  end_row=end_row, end_column=measurement_idx)
                    ws.merge_cells(start_row=start_row, start_column=gw_idx,
                                  end_row=end_row, end_column=gw_idx)
                    ws.merge_cells(start_row=start_row, start_column=carton_no_idx,
                                  end_row=end_row, end_column=carton_no_idx)
                
                # 开始新组
                current_carton = carton_no
                start_row = row_idx
        
        # Handle the last group before the Total row
        if start_row and start_row < effective_last_row:  # More than one row in the last group
            end_row = effective_last_row
            ws.merge_cells(start_row=start_row, start_column=ctns_idx, 
                          end_row=end_row, end_column=ctns_idx)
            ws.merge_cells(start_row=start_row, start_column=measurement_idx,
                          end_row=end_row, end_column=measurement_idx)
            ws.merge_cells(start_row=start_row, start_column=gw_idx,
                          end_row=end_row, end_column=gw_idx)
            ws.merge_cells(start_row=start_row, start_column=carton_no_idx,
                          end_row=end_row, end_column=carton_no_idx)
        
        # Set vertical alignment for merged cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.coordinate in ws.merged_cells:
                    cell.alignment = Alignment(vertical='center')
        
        wb.save(workbook_path)
        print(f"Successfully merged cells in Packing List for {workbook_path}")
        return True
    except Exception as e:
        print(f"Error merging cells in Packing List: {e}")
        return False

def apply_pl_footer_styling(workbook_path):
    """为PL页脚应用样式，包括合并单元格和加粗文本。"""
    try:
        wb = load_workbook(workbook_path)
        sheet_name = 'PL' if 'PL' in wb.sheetnames else 'Packing List'
        
        if sheet_name not in wb.sheetnames:
            print(f"No '{sheet_name}' sheet found in workbook")
            return False
            
        ws = wb[sheet_name]
        
        # 查找总计行和页脚行
        total_row = None
        footer_start_row = None
        
        for row_idx in range(1, ws.max_row + 1):
            if ws.cell(row=row_idx, column=3).value == 'Total':  # 假设第三列是'名称'/'DESCRIPTION'列
                total_row = row_idx
            elif ws.cell(row=row_idx, column=1).value and 'PACKED IN' in str(ws.cell(row=row_idx, column=1).value):
                footer_start_row = row_idx
                break
        
        if not footer_start_row:
            # 尝试完整文本匹配
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value and 'PACKED IN' in str(cell_value):
                    footer_start_row = row_idx
                    break
        
        if not footer_start_row:
            print("Footer rows not found in the sheet")
            return False
        
        # 获取列数
        max_column = ws.max_column
        
        # 绿色背景填充
        light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # 对每个页脚行应用样式
        for row_idx in range(footer_start_row, footer_start_row + 5):  # 假设有5行页脚
            if row_idx <= ws.max_row:
                # 合并整行所有单元格
                try:
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_column)
                except Exception as e:
                    print(f"Warning: Could not merge cells for row {row_idx}: {e}")
                
                # 为每个单元格应用样式
                cell = ws.cell(row=row_idx, column=1)
                
                # 设置加粗字体
                cell.font = Font(bold=True)
                
                # 设置绿色背景
                cell.fill = light_green_fill
                
                # 设置左对齐
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 设置边框
                thin_border = Side(style='thin')
                cell.border = Border(
                    left=thin_border,
                    right=thin_border,
                    top=thin_border,
                    bottom=thin_border
                )
        
        # 保存样式更改
        wb.save(workbook_path)
        print(f"Successfully applied footer styling to {sheet_name} sheet")
        return True
        
    except Exception as e:
        print(f"Error applying footer styling: {e}")
        return False

# Function to apply styling to Excel workbook
def apply_excel_styling(file_path):
    """Apply professional styling to the Excel workbook."""
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Column widths
    column_widths = {
        'NO.': 6,
        'Material code': 20,
        'DESCRIPTION': 30,
        'Model NO.': 20,
        'Unit Price': 12,
        'Qty': 8,
        'Unit': 8,
        'Amount': 12,
        'net weight': 12,
        '采购单价': 12,
        '采购总价': 12,
        'FOB单价': 12,
        'FOB总价': 12,
        '总保费': 10,
        '总运费': 10,
        '每公斤摊的运保费': 18,
        '该项对应的运保费': 18,
        'CIF总价(FOB总价+运保费)': 20,
        'CIF单价': 12,
        '单价USD数值': 12,
        '单位': 8,
        '开票品名': 15,
        'factory': 12,
        'project': 15,
        'end use': 15
    }
    
    # Apply header styling
    for col_idx, col in enumerate(ws[1], 1):
        col.font = header_font
        col.fill = header_fill
        col.alignment = header_alignment
        
        # Set column width if specified
        col_name = ws.cell(row=1, column=col_idx).value
        if col_name in column_widths:
            ws.column_dimensions[get_column_letter(col_idx)].width = column_widths[col_name]
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    # Apply data cell styling
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        for col_idx, cell in enumerate(row, 1):
            col_name = ws.cell(row=1, column=col_idx).value
            
            # Apply cell styling based on column type
            if col_name in ['net weight','FOB单价','Amount', 'FOB总价', 'CIF单价', 'CIF总价(FOB总价+运保费)', '采购单价', '采购总价', '总保费', '总运费', '每公斤摊的运保费', '该项对应的运保费', '单价USD数值']:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            elif col_name in ['Qty' ]:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col_name in ['Unit Price' ]:
                cell.number_format = '#,##0.0000'
                cell.alignment = Alignment(horizontal='right')
            elif col_name in ['NO.']:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Add borders to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the styled workbook
    wb.save(file_path)

# Function to add summary row to Excel file
def add_summary_row(df, file_path):
    """Add a summary row to the Excel file and save."""
    # Calculate totals for numeric columns
    numeric_cols = ['Qty', 'net weight', '采购总价', 'FOB总价', 'CIF总价(FOB总价+运保费)', 'Amount']
    
    # Create a summary row
    summary = {}
    summary['Material code'] = 'Total'
    
    for col in numeric_cols:
        if col in df.columns:
            summary[col] = df[col].sum()
    
    # Append summary row to DataFrame
    summary_df = pd.DataFrame([summary])
    df_with_summary = pd.concat([df, summary_df], ignore_index=True)
    
    # Write to Excel with retries to handle permission issues
    max_retries = 3
    retry_delay = 2  # seconds
    
    for attempt in range(max_retries):
        try:
            # Safe save to handle file access issues
            df_with_summary.to_excel(file_path, index=False)
            # If we get here, the save was successful
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                print(f"File {file_path} is locked. Retrying in {retry_delay} seconds... (Attempt {attempt+1}/{max_retries})")
                time.sleep(retry_delay)
            else:
                print(f"Could not save to {file_path} after {max_retries} attempts due to permission issues.")
                print("Please close any applications that might have this file open.")
                raise
        except Exception as e:
            print(f"Unexpected error while saving {file_path}: {e}")
            raise

# Function to safely save DataFrame to Excel
def safe_save_to_excel(df, file_path, include_summary=True):
    """Safely save DataFrame to Excel with proper error handling."""
    try:
        if include_summary:
            # Add a summary row
            add_summary_row(df, file_path)
        else:
            # Save directly without summary
            df.to_excel(file_path, index=False)
            
        # Apply styling
        try:
            apply_excel_styling(file_path)
            print(f"Successfully saved and styled: {file_path}")
            return True
        except Exception as e:
            print(f"Warning: File saved but could not apply styling: {e}")
            return True
    except PermissionError as e:
        print(f"Error: Could not save to {file_path} due to permission issues.")
        print(f"Please close the file if it's open in another application.")
        print(f"Error details: {e}")
        return False
    except Exception as e:
        print(f"Error: Failed to save to {file_path}.")
        print(f"Error details: {e}")
        return False

# Helper function to find columns with specific patterns
def find_column_with_pattern(df, patterns, target_col_name=None):
    """Find a column that contains any of the given patterns."""
    for col in df.columns:
        col_str = str(col).lower()
        
        # For multi-level headers that were combined with separator
        if '|' in col_str:
            # Split the combined name to check both English and Chinese parts
            parts = col_str.split('|')
            for pattern in patterns:
                pattern_lower = pattern.lower()
                # Check if pattern matches any part of the column name
                if any(pattern_lower in part for part in parts):
                    if target_col_name:
                        print(f"Found column '{col}' for {target_col_name}")
                    return col
        else:
            # Regular column name check
            for pattern in patterns:
                if pattern.lower() in col_str:
                    if target_col_name:
                        print(f"Found column '{col}' for {target_col_name}")
                    return col
    
    if target_col_name:
        print(f"WARNING: Could not find a column matching patterns {patterns} for {target_col_name}")
    return None

# Helper function to print found mappings
def print_column_mappings(mappings):
    """Print a summary of all column mappings found."""
    print("\nColumn mappings summary:")
    missing_cols = []
    
    # Define the expected column list
    expected_columns = [
        'NO.', 'Material code', 'DESCRIPTION', 'Model NO.', 'Unit Price', 'Qty', 'Unit', 
        'net weight', 'factory', 'project', 'end use'
    ]

    pack_list_expected_columns = [
        'Sr No.',
        'P/N.',
        'DESCRIPTION', 
        'Model NO.',
        'QUANTITY',
        'CTNS',
        'Carton MEASUREMENT',
        'G.W (KG)',
        'N.W(KG)',
        'Carton NO.'
    ]

    
    # Print the found mappings
    print("\nFound column mappings:")
    for target_col in expected_columns:
        if target_col in mappings:
            print(f"  {target_col} <- {mappings[target_col]}")
        else:
            missing_cols.append(target_col)
    
    # Print missing mappings
    if missing_cols:
        print("\nMissing column mappings:")
        for col in missing_cols:
            print(f"  {col} - No matching column found in the source file")
    
    # Summary
    found_count = len(mappings)
    expected_count = len(expected_columns)
    print(f"\nFound {found_count} out of {expected_count} expected column mappings ({100*found_count/expected_count:.1f}%)")

def split_by_project_and_factory(df):
    """Split the dataframe by project and factory."""
    print("Available columns for splitting:", df.columns.tolist())
    
    # 确保必要的列存在
    if 'project' not in df.columns:
        print("WARNING: 'project'列不存在，添加默认值'大华'")
        df['project'] = '大华'
    
    if 'factory' not in df.columns:
        print("WARNING: 'factory'列不存在，添加默认值'默认工厂'")
        df['factory'] = '默认工厂'
    
    # Clean up project and factory values
    # Convert NaN, None and empty strings to default values
    df['project'] = df['project'].apply(lambda x: '大华' if pd.isna(x) or str(x).strip() == '' else str(x).strip())
    df['factory'] = df['factory'].apply(lambda x: '默认工厂' if pd.isna(x) or str(x).strip() == '' else str(x).strip())
    
    print("Unique project values:", df['project'].unique())
    print("Unique factory values:", df['factory'].unique())
    
    # Define the project categories with more robust string handling
    project_categories = {
        '大华': lambda x: str(x).strip() == '大华',
        '麦格米特': lambda x: str(x).strip() == '麦格米特',
        '工厂': lambda x: str(x).strip() not in ['大华', '麦格米特']
    }
    
    # Get unique factories
    factories = sorted(df['factory'].unique())
    if len(factories) == 0:
        factories = ['默认工厂']
        df['factory'] = '默认工厂'
        print("WARNING: 没有有效的工厂值，使用'默认工厂'")
    
    # Dictionary to store split dataframes
    split_dfs = {}
    
    # Split by project and factory
    for project_name, project_filter in project_categories.items():
        try:
            project_df = df[df['project'].apply(project_filter)]
            print(f"Found {len(project_df)} rows for project {project_name}")
            
            for factory in factories:
                key = (project_name, factory)
                factory_df = project_df[project_df['factory'] == factory]
                
                split_dfs[key] = factory_df
                print(f"Found {len(split_dfs[key])} rows for {project_name} - {factory}")
        except Exception as e:
            print(f"Error processing project {project_name}: {e}")
            continue
    
    # 如果有项目为'工厂'的数据集为空，确保仍添加一个空的DataFrame
    for factory in factories:
        key = ('工厂', factory)
        if key not in split_dfs:
            split_dfs[key] = pd.DataFrame(columns=df.columns)
            print(f"Added empty DataFrame for 工厂 - {factory}")
    
    # 检查是否所有项目和工厂的组合都有数据框
    all_projects = ['大华', '麦格米特', '工厂']
    for project in all_projects:
        for factory in factories:
            key = (project, factory)
            if key not in split_dfs:
                split_dfs[key] = pd.DataFrame(columns=df.columns)
                print(f"Added empty DataFrame for {project} - {factory}")
    
    return split_dfs, project_categories

# Function to generate valid invoice sheet name
def generate_invoice_sheet_name(prefix="CXCI"):
    """Generate a valid invoice sheet name in the format XXXX20230101####"""
    import datetime
    today = datetime.datetime.now()
    date_part = today.strftime("%Y%m%d")

    
    import datetime

# 获取当前时间
    now = datetime.datetime.now()
    # 获取当天 0 点的时间
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    # 计算当前时间与当天 0 点的时间差
    delta = now - midnight
    # 获取时间差的总秒数
    seconds = delta.total_seconds()

    serial = int((seconds / 86400) * 10000)

    # Use a serial number starting with 0001
    return f"{prefix}{date_part}{serial}"

# Function to convert numbers to English words
def num_to_words(num):
    """Convert a number to its English word representation."""
    ones = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 'NINE', 'TEN', 
            'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN', 'SIXTEEN', 'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
    tens = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY', 'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']
    
    def _convert_less_than_thousand(num):
        if num < 20:
            return ones[num]
        elif num < 100:
            return tens[num // 10] + ('-' + ones[num % 10] if num % 10 > 0 else '')
        else:
            return ones[num // 100] + ' HUNDRED' + (' AND ' + _convert_less_than_thousand(num % 100) if num % 100 > 0 else '')
    
    if num == 0:
        return 'ZERO'
    
    integer_part = int(num)
    decimal_part = int(round((num - integer_part) * 100))
    
    result = ''
    
    if integer_part >= 1000000000:
        result += _convert_less_than_thousand(integer_part // 1000000000) + ' BILLION'
        integer_part %= 1000000000
        if integer_part > 0:
            result += ' '
    
    if integer_part >= 1000000:
        result += _convert_less_than_thousand(integer_part // 1000000) + ' MILLION'
        integer_part %= 1000000
        if integer_part > 0:
            result += ' '
    
    if integer_part >= 1000:
        result += _convert_less_than_thousand(integer_part // 1000) + ' THOUSAND'
        integer_part %= 1000
        if integer_part > 0:
            result += ' '
            if integer_part < 100:  # Add 'AND' for values less than 100
                result += 'AND '
    
    if integer_part > 0:
        result += _convert_less_than_thousand(integer_part)
    
    # Handle decimal part if present
    if decimal_part > 0:
        result += ' AND CENTS ' + (tens[decimal_part // 10] + (('-' + ones[decimal_part % 10]) if decimal_part % 10 > 0 else '') if decimal_part >= 20 else ones[decimal_part])
    
    return result.strip()

def modify_header_file(h_file_path, company_name, company_address):
    """
    Modify the header file (h.xlsx) with company information while preserving styles and merges.
    
    Args:
        h_file_path: Path to the h.xlsx file
        company_name: Company name to replace A1
        company_address: Company address to replace A2
    """
    try:
        # Load the workbook
        wb = load_workbook(h_file_path)
        ws = wb.active
        
        # Store the original merged cell ranges
        merged_ranges = list(ws.merged_cells.ranges)
        
        # Unmerge all cells temporarily
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
        
        # Update the values
        ws['A1'] = company_name
        ws['A2'] = company_address
        
        # Reapply the merges
        for merged_range in merged_ranges:
            ws.merge_cells(str(merged_range))
        
        # Ensure text alignment is preserved
        for row in [1, 2]:
            cell = ws.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Save the modified file
        wb.save(h_file_path)
        print(f"Successfully updated header file with company information")
        return True
    except Exception as e:
        print(f"Error modifying header file: {e}")
        return False

def read_policy_file(policy_file):
    """
    读取新格式的政策文件并提取所需参数
    
    Args:
        policy_file: 政策文件路径
    
    Returns:
        dict: 包含所有政策参数的字典
    """
    try:
        # 读取政策文件，使用第一列作为索引
        policy_df = pd.read_excel(policy_file, index_col=0)
        
        # 提取所需参数
        # 使用字段名称作为索引来获取对应的值
        policy_params = {
            'packing_list_no': policy_df.loc['采购装箱单编号', '值'],  # 装箱单编号
            'total_net_weight': float(policy_df.loc['总净重(KG)', '值']),  # 总净重
            'total_freight': float(policy_df.loc['总运费(RMB)', '值']),  # 总运费
            'freight_unit_price': float(policy_df.loc['运费单价(RMB/KG)', '值']),  # 运费单价
            'markup_percentage': float(policy_df.loc['加价率', '值']),  # 加价率
            'insurance_coefficient': float(policy_df.loc['保险系数', '值']),  # 保险系数
            'insurance_rate': float(policy_df.loc['保险费率', '值']),  # 保险费率
            'exchange_rate': float(policy_df.loc['汇率(RMB/美元)', '值']),  # 汇率
            'company_name': policy_df.loc['公司名称', '值'],  # 公司名称
            'company_address': policy_df.loc['公司地址', '值'],  # 公司地址
            'bank_account': policy_df.loc['Account number', '值'],  # 银行账号
            'bank_name': policy_df.loc['Bank Name', '值'],  # 银行名称
            'bank_address': policy_df.loc['Bank Address', '值'],  # 银行地址
            'swift_no': policy_df.loc['SWIFT No.', '值']  # SWIFT号码
        }
        
        # 打印提取的参数用于调试
        print("\n提取的政策参数:")
        for key, value in policy_params.items():
            print(f"{key}: {value}")
        
        return policy_params
    except Exception as e:
        print(f"读取政策文件时出错: {e}")
        print("政策文件的实际内容:")
        try:
            temp_df = pd.read_excel(policy_file)
            print(temp_df.head())
            print("\n列名:", temp_df.columns.tolist())
        except Exception as e2:
            print(f"无法读取政策文件内容: {e2}")
        raise

# Main function to process the shipping list
def process_shipping_list(packing_list_file, policy_file, output_dir='outputs'):
    # Read the input files
    packing_list_df = read_excel_file(packing_list_file, skip=2)
    
    # 使用新的政策文件读取函数
    try:
        policy_params = read_policy_file(policy_file)
        
        # 从政策参数中提取值
        markup_percentage = policy_params['markup_percentage']  # 加价率
        insurance_coefficient = policy_params['insurance_coefficient']  # 保险系数
        insurance_rate = policy_params['insurance_rate']  # 保险费率
        total_freight_amount = policy_params['total_freight']  # 总运费
        exchange_rate = policy_params['exchange_rate']  # 汇率
        pc = policy_params['company_name']  # 公司名称
        pca = policy_params['company_address']  # 公司地址
        ba = policy_params['bank_account']  # 银行账号
        bn = policy_params['bank_name']  # 银行名称
        badd = policy_params['bank_address']  # 银行地址
        swn = policy_params['swift_no']  # SWIFT号码
        
    except Exception as e:
        print(f"处理政策文件时出错: {e}")
        raise

    # Modify h.xlsx with company information
    h_file = 'h.xlsx'
    h_file_paths = [
        h_file,  # Current directory
        os.path.join(output_dir, h_file),  # Output directory
        os.path.join(os.path.dirname(os.path.abspath(__file__)), h_file)  # Script directory
    ]
    
    h_file_found = False
    for h_path in h_file_paths:
        if os.path.exists(h_path):
            print(f"Found h.xlsx at: {h_path}")
            if modify_header_file(h_path, pc, pca):
                h_file_found = True
                break
    
    if not h_file_found:
        print("Warning: Could not find or modify h.xlsx")

    # Update pl_h.xlsx with company information
    pl_h_file = 'pl_h.xlsx'
    try:
        # Load the workbook
        wb = load_workbook(pl_h_file)
        ws = wb.active
        
        # Store the original merged cell ranges
        merged_ranges = list(ws.merged_cells.ranges)
        
        # Unmerge cells temporarily
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
        
        # Update the values in B4 and B5 (shipper information)
        ws['B4'] = pc  # Company name
        ws['B5'] = pca  # Company address

        ws['A1'] = pc  # Company name
        ws['A2'] = pca  # Company address
        
        # Reapply the merges
        for merged_range in merged_ranges:
            ws.merge_cells(str(merged_range))
        
        # Save the modified file
        wb.save(pl_h_file)
        print(f"Successfully updated pl_h.xlsx with company information")
    except Exception as e:
        print(f"Error modifying pl_h.xlsx: {e}")

    # Print original column names for debugging
    print("Original packing list columns:")
    for col in packing_list_df.columns:
        print(f"  {col}")
    
    # 打印前10行数据，验证是否正确读取了所有行
    print("\nVerifying first 10 rows of data:")
    preview_rows = min(10, len(packing_list_df))
    for i in range(preview_rows):
        first_col_value = packing_list_df.iloc[i, 0] if not packing_list_df.empty and len(packing_list_df.columns) > 0 else "N/A"
        print(f"  Row {i+1}: {first_col_value}")
    
    # Define packing list output columns - define this at the beginning so it's available everywhere
    pl_output_columns = [
        'S/N', 'Part Number', '名称', 'Model Number', 'Quantity', 
        'Total Carton Quantity', 'Total Volume (CBM)', 'Total Gross Weight (kg)', 
        'Total Net Weight (kg)', 'Carton Number'
    ]
    
    print("\nPacking List output columns defined:")
    print(pl_output_columns)
    
    # Clean up the column names for better handling
    packing_list_df.columns = [str(col).strip() for col in packing_list_df.columns]
    
    # Create new DataFrames for the processed data
    result_df = pd.DataFrame()
    pl_result_df = pd.DataFrame()
    
    # Keep track of mappings for debugging
    column_mappings = {}
    
    # Find key columns by pattern matching
    print("\nFinding column mappings...")
    # Main invoice columns
    sr_no_col = find_column_with_pattern(packing_list_df, ['S/N', '序号', '序列号'], 'NO.')
    material_code_col = find_column_with_pattern(packing_list_df, ['p/n', 'Part Number', 'material code', '系统料号', '料号'], 'Material code')
    # 修改: 将 '供应商开票名称' 放在匹配模式的最前面，优先使用该字段作为DESCRIPTION
    description_col = find_column_with_pattern(packing_list_df, ['供应商开票名称', 'Commercial Invoice Description', '清关英文货描(关务提供)', '描述', 'description'], 'DESCRIPTION')
    # 新增：查找进口清关货描（Commodity Description (Customs)）列
    customs_desc_col = find_column_with_pattern(packing_list_df, ['进口清关货描', 'Commodity Description (Customs)'], 'Commodity Description (Customs)')
    model_col = find_column_with_pattern(packing_list_df, ['Model Number', '型号', '物料型号', '货物型号', 'model'], 'Model NO.')
    unit_price_col = find_column_with_pattern(packing_list_df, ['Unit Price (Excl. Tax, CNY)()', 'unit price', '采购单价不含税', '不含税单价', '单价'], 'Unit Price')
    qty_col = find_column_with_pattern(packing_list_df, ['Quantity', 'quantity', '数量', 'qty'], 'Qty')
    unit_col = find_column_with_pattern(packing_list_df, ['Unit', '单位', '单位中文'], 'Unit')
    
    # Enhanced patterns for net weight and gross weight
    # 优先级顺序很重要 - 总净重相关的模式必须放在最前面
    total_net_weight_patterns = [
        'Total Net Weight (kg)',  # 最优先匹配
        'Total Net Weight',
        'N.W  (KG)总净重',
        '总净重(KG)',
        '总净重',
        'Total N.W'
    ]
    
    # 单件净重的排除模式
    unit_net_weight_patterns = [
        'Net Weight per Unit',
        'Unit Net Weight',
        'per unit',
        'per piece',
        'unit net',
        '单件净重',
        '每件净重',
        '单个净重'
    ]
    
    # 其他净重相关的模式（仅在找不到总净重时使用）
    general_net_weight_patterns = [
        'N.W(KG)',
        'N.W  (KG)',
        'net weight',
        'n/w',
        '净重',
        'Net Weight',
        'net wt'
    ]
    
    def find_total_net_weight_column(df, total_patterns, unit_patterns, general_patterns):
        """专门用于查找总净重列的函数"""
        # 1. 首先尝试精确匹配总净重模式
        for pattern in total_patterns:
            for col in df.columns:
                col_str = str(col).lower()
                pattern_lower = pattern.lower()
                if pattern_lower in col_str:
                    # 确保这不是单件净重列
                    if not any(unit_p.lower() in col_str for unit_p in unit_patterns):
                        print(f"Found exact total net weight match: {col}")
                        return col
        
        # 2. 如果没找到精确匹配，检查其他净重列，但要排除单件净重
        for pattern in general_patterns:
            for col in df.columns:
                col_str = str(col).lower()
                pattern_lower = pattern.lower()
                if pattern_lower in col_str:
                    # 再次确认不是单件净重列
                    if not any(unit_p.lower() in col_str for unit_p in unit_patterns):
                        print(f"Found general net weight match: {col}")
                        return col
        
        print("WARNING: Could not find appropriate total net weight column")
        return None
    
    # 使用新的查找函数找到总净重列
    net_weight_col = find_total_net_weight_column(
        packing_list_df,
        total_net_weight_patterns,
        unit_net_weight_patterns,
        general_net_weight_patterns
    )
    
    if net_weight_col:
        print(f"Selected net weight column for both invoice and packing list: {net_weight_col}")
        # 为主发票设置净重列
        result_df['net weight'] = packing_list_df[net_weight_col]
        result_df['Total Net Weight (kg)'] = packing_list_df[net_weight_col]
        # 为装箱单设置净重列
        pl_result_df['N.W(KG)'] = packing_list_df[net_weight_col]
        
        column_mappings['net weight'] = net_weight_col
        column_mappings['Total Net Weight (kg)'] = net_weight_col
        print(f"Using '{net_weight_col}' for all net weight values")
    else:
        print("ERROR: Could not find total net weight column, using fallback values")
        # 设置默认值
        result_df['net weight'] = 0
        result_df['Total Net Weight (kg)'] = 0
        pl_result_df['N.W(KG)'] = 0
    
    # 修改总毛重匹配模式
    gross_weight_patterns = [
        'Total Gross Weight (kg)',  # 优先匹配总毛重
        'Total Gross Weight',
        '总毛重',
        'G.W  (KG)总毛重', 
        'Total G.W',
        'gross weight', 
        'g.w', 
        'G.W (KG)', 
        'G.W  (KG)'
    ]
    
    # 在find_column_with_pattern函数调用时使用新的匹配模式
    gross_weight_col = find_column_with_pattern(packing_list_df, gross_weight_patterns, 'G.W (KG)')
    
    # 扩展工厂列的匹配模式
    factory_patterns = [
        'Plant Location', 
        '工厂', 
        'factory', 
        'daman/silvass', 
        '工厂地点', 
        '送达方',
        '目的地',
        '送货地点',
        'location',
        'delivery location',
        'plant',
        '厂区'
    ]
    factory_col = find_column_with_pattern(packing_list_df, factory_patterns, 'factory')
    
    # 如果工厂列未找到，创建一个默认值
    if factory_col is None:
        print("WARNING: 未找到工厂列，使用默认值'默认工厂'")
        packing_list_df['默认工厂'] = '默认工厂'
        factory_col = '默认工厂'
    
    # 扩展项目列的匹配模式
    project_patterns = [
        'Project', 
        '项目名称', 
        '项目', 
        'project name',
        'program',
        'program name',
        '计划名称',
        '方案名称',
        '所属项目'
    ]
    project_col = find_column_with_pattern(packing_list_df, project_patterns, 'project')
    
    # 如果项目列未找到，创建一个默认值
    if project_col is None:
        print("WARNING: 未找到项目列，使用默认值'大华'")
        packing_list_df['默认项目'] = '大华'
        project_col = '默认项目'
        
    end_use_col = find_column_with_pattern(packing_list_df, ['end use', '用途'], 'end use')
    
    # Additional packing list columns
    ctns_col = find_column_with_pattern(packing_list_df, ['ctns', '件数'], 'CTNS')
    # find_column_with_pattern takes 3 parameters:
    # 1. packing_list_df: The dataframe to search in
    # 2. ['体积（CBM）', '总体积', 'measurement']: Array of possible column names to look for in Chinese/English
    #    - Will match any column that contains these strings
    # 3. 'Carton MEASUREMENT': The standardized name to use if a match is found
    #    - This is what the matched column will be renamed to in the output
    carton_measurement_col = find_column_with_pattern(packing_list_df, ['体积（CBM）', '总体积', 'CBM'], 'Carton MEASUREMENT')
    carton_no_col = find_column_with_pattern(packing_list_df, ['carton no', '箱号', 'ctn no'], 'Carton NO.')
    
    # 贸易类型列
    trade_type_col = find_column_with_pattern(packing_list_df, ['出口报关方式', '贸易方式', 'trade type'], 'Trade Type')
    
    # Map main invoice columns
    if sr_no_col:
        result_df['NO.'] = packing_list_df[sr_no_col]
        column_mappings['NO.'] = sr_no_col
    else:
        # Create a default sequence number
        result_df['NO.'] = range(1, len(packing_list_df) + 1)
        print("WARNING: S/N column not found, generating sequence numbers")
    
    if material_code_col:
        result_df['Material code'] = packing_list_df[material_code_col]
        column_mappings['Material code'] = material_code_col
    else:
        # Material code is essential - set to empty if not found
        result_df['Material code'] = ""
        print("WARNING: Material code column not found, using empty values")
    
    if description_col:
        result_df['DESCRIPTION'] = packing_list_df[description_col]
        column_mappings['DESCRIPTION'] = description_col
        # 添加更明确的日志消息，区分不同的列来源
        if '供应商开票名称' in str(description_col):
            print(f"Using '供应商开票名称' column '{description_col}' for DESCRIPTION as recommended")
        else:
            print(f"Using '{description_col}' as the source for DESCRIPTION")
    else:
        # Description is essential - default to Material code if not found
        if material_code_col:
            result_df['DESCRIPTION'] = packing_list_df[material_code_col]
            print("WARNING: Description column not found, using Material code as Description")
        else:
            result_df['DESCRIPTION'] = "Unknown Material"
            print("WARNING: Description and Material code columns not found")
    
    if model_col:
        result_df['Model NO.'] = packing_list_df[model_col]
        column_mappings['Model NO.'] = model_col
    else:
        # Default to empty model number
        result_df['Model NO.'] = ""
        print("WARNING: Model Number column not found")
    
    if unit_price_col:
        result_df['Unit Price'] = packing_list_df[unit_price_col]
        column_mappings['Unit Price'] = unit_price_col
    else:
        # Unit price is essential for calculations - set default
        result_df['Unit Price'] = 0
        print("WARNING: Unit Price column not found, using zeros")
    
    if qty_col:
        result_df['Qty'] = packing_list_df[qty_col]
        column_mappings['Qty'] = qty_col
    else:
        # Quantity is essential - set default
        result_df['Qty'] = 1
        print("WARNING: Quantity column not found, using default of 1")
    
    if unit_col:
        # First, copy the original units to both dataframes
        result_df['Unit'] = packing_list_df[unit_col]
        result_df['Original_Unit'] = packing_list_df[unit_col]  # Keep original units for reference
        pl_result_df['Original_Unit'] = packing_list_df[unit_col]  # Also keep in pl_result_df
        column_mappings['Unit'] = unit_col
    
    if gross_weight_col:
        print(f"Using total gross weight column: {gross_weight_col} for G.W (KG)")
        pl_result_df['G.W (KG)'] = packing_list_df[gross_weight_col]
    else:
        # 如果找不到总毛重列，尝试计算
        unit_gw_col = find_column_with_pattern(packing_list_df, ['Gross Weight per Unit', '单件毛重'], 'Unit Gross Weight')
        if unit_gw_col and 'QUANTITY' in pl_result_df.columns:
            print(f"Calculating total gross weight from unit weight and quantity")
            pl_result_df['G.W (KG)'] = pd.to_numeric(packing_list_df[unit_gw_col], errors='coerce') * pd.to_numeric(pl_result_df['QUANTITY'], errors='coerce')
        else:
            print("WARNING: Could not find or calculate total gross weight")
            pl_result_df['G.W (KG)'] = None
    
    if factory_col:
        result_df['factory'] = packing_list_df[factory_col]
        column_mappings['factory'] = factory_col
    else:
        # 设置默认工厂值
        result_df['factory'] = '默认工厂'
        print("WARNING: 工厂列未找到，添加默认值'默认工厂'")
    
    if project_col:
        result_df['project'] = packing_list_df[project_col]
        column_mappings['project'] = project_col
    else:
        # 设置默认项目值
        result_df['project'] = '大华'
        print("WARNING: 项目列未找到，添加默认值'大华'")
    
    if end_use_col:
        result_df['end use'] = packing_list_df[end_use_col]
        column_mappings['end use'] = end_use_col
    
    # Map packing list columns
    if sr_no_col:
        pl_result_df['S/N'] = packing_list_df[sr_no_col]
    
    if material_code_col:
        pl_result_df['Part Number'] = packing_list_df[material_code_col]
    
    if description_col:
        pl_result_df['名称'] = packing_list_df[description_col]
        # 添加日志，显示使用的是相同的描述源
        if '供应商开票名称' in str(description_col):
            print(f"Using '供应商开票名称' column '{description_col}' for packing list 名称 as well")
    
    if model_col:
        pl_result_df['Model Number'] = packing_list_df[model_col]
    
    if qty_col:
        pl_result_df['Quantity'] = packing_list_df[qty_col]
    
    if ctns_col:
        pl_result_df['Total Carton Quantity'] = packing_list_df[ctns_col]
    else:
        pl_result_df['Total Carton Quantity'] = 1  # Default value if not found
    
    if carton_measurement_col:
        pl_result_df['Total Volume (CBM)'] = packing_list_df[carton_measurement_col]
    else:
        pl_result_df['Total Volume (CBM)'] = ""  # Default empty value if not found
    
    if gross_weight_col:
        pl_result_df['Total Gross Weight (kg)'] = packing_list_df[gross_weight_col]
    else:
        pl_result_df['Total Gross Weight (kg)'] = ""  # Default empty value if not found
    
    if net_weight_col:
        pl_result_df['Total Net Weight (kg)'] = packing_list_df[net_weight_col]
    else:
        pl_result_df['Total Net Weight (kg)'] = 0  # Default value if not found
    
    if carton_no_col:
        pl_result_df['Carton Number'] = packing_list_df[carton_no_col]
    else:
        pl_result_df['Carton Number'] = ""  # Default empty value if not found
    
    # Add Trade Type column
    if trade_type_col:
        result_df['Trade Type'] = packing_list_df[trade_type_col]
        pl_result_df['Trade Type'] = packing_list_df[trade_type_col]
        column_mappings['Trade Type'] = trade_type_col
    else:
        # Try to find 出口报关方式 column
        report_type_col = find_column_with_pattern(packing_list_df, ['出口报关方式'], '出口报关方式')
        if report_type_col:
            result_df['Trade Type'] = packing_list_df[report_type_col]
            pl_result_df['Trade Type'] = packing_list_df[report_type_col]
            column_mappings['Trade Type'] = report_type_col
        else:
            print("WARNING: 无法确定贸易类型，默认将所有物料视为一般贸易处理")
            result_df['Trade Type'] = '一般贸易'  # Default to general trade
            pl_result_df['Trade Type'] = '一般贸易'  # Default to general trade
    
    # Add factory to packing list
    if factory_col:
        pl_result_df['factory'] = packing_list_df[factory_col]
    else:
        pl_result_df['factory'] = '默认工厂'
        print("WARNING: 为打包清单设置默认工厂值'默认工厂'")
    
    # Map project column directly from source
    if project_col and project_col in packing_list_df.columns:
        result_df['project'] = packing_list_df[project_col]
        pl_result_df['project'] = packing_list_df[project_col]
        column_mappings['project'] = project_col
        print(f"Successfully mapped project column from '{project_col}'")
    elif '项目名称' in packing_list_df.columns:
        result_df['project'] = packing_list_df['项目名称']
        pl_result_df['project'] = packing_list_df['项目名称']
        column_mappings['project'] = '项目名称'
        print(f"Successfully mapped project column from '项目名称'")
    else:
        print(f"Warning: Project column not found in packing list, using default value")
        result_df['project'] = '大华'  # 使用大华作为默认项目
        pl_result_df['project'] = '大华'  # 确保打包清单也有相同的项目值
    
    # Print found mappings for debugging
    print_column_mappings(column_mappings)
    
    # Debug print for net weight column specifically
    if net_weight_col:
        print(f"\nFound net weight column: {net_weight_col}")
        # Print some sample values
        print("Sample net weight values:")
        for i, val in enumerate(packing_list_df[net_weight_col].head(5)):
            print(f"  Row {i+1}: {val}")
    else:
        print("\nWARNING: Net weight column not found!")
        print("Available columns:")
        for col in packing_list_df.columns:
            print(f"  {col}")
    
    # Apply trade type determination
    def determine_trade_type(row_type):
        if pd.isna(row_type):
            return '一般贸易'  # Default to general trade if empty
        
        row_type_str = str(row_type).strip().lower()
        if '买单' in row_type_str:
            return '买单贸易'
        else:
            return '一般贸易'
    
    # Apply trade type determination to both DataFrames
    result_df['Trade Type'] = result_df['Trade Type'].apply(determine_trade_type)
    pl_result_df['Trade Type'] = pl_result_df['Trade Type'].apply(determine_trade_type)
    
    # Count items by trade type
    general_trade_count = (result_df['Trade Type'] == '一般贸易').sum()
    purchase_trade_count = (result_df['Trade Type'] == '买单贸易').sum()
    print(f"\n贸易类型统计：")
    print(f"  一般贸易物料数量: {general_trade_count}")
    print(f"  买单贸易物料数量: {purchase_trade_count}")
    
    # Set Shipper information for both DataFrames
    result_df['Shipper'] = result_df['Trade Type'].apply(
        lambda x: '创想(创想-PCT)' if x == '一般贸易' else 'Unicair(UC-PCT)'
    )
    
    pl_result_df['Shipper'] = pl_result_df['Trade Type'].apply(
        lambda x: '创想(创想-PCT)' if x == '一般贸易' else 'Unicair(UC-PCT)'
    )
    
    # If Amount column is missing, set to None
    if 'Amount' not in result_df.columns:
        result_df['Amount'] = None
        
    # Convert columns to numeric for calculations
    try:
        if net_weight_col:
            result_df['net weight'] = pd.to_numeric(result_df['net weight'], errors='coerce')
            result_df['Total Net Weight (kg)'] = pd.to_numeric(result_df['Total Net Weight (kg)'], errors='coerce')
            print(f"Converted net weight to numeric. Example values: {result_df['net weight'].head()}")
        else:
            # If net weight column not found, set to default value
            print("WARNING: Setting default values for missing net weight column")
            result_df['net weight'] = 0
            result_df['Total Net Weight (kg)'] = 0
            
        result_df['Qty'] = pd.to_numeric(result_df['Qty'], errors='coerce')
        result_df['Unit Price'] = pd.to_numeric(result_df['Unit Price'], errors='coerce')
        
        # Also convert packing list numeric columns
        pl_result_df['Quantity'] = pd.to_numeric(pl_result_df['Quantity'], errors='coerce')
        if 'Total Gross Weight (kg)' in pl_result_df.columns:
            pl_result_df['Total Gross Weight (kg)'] = pd.to_numeric(pl_result_df['Total Gross Weight (kg)'], errors='coerce')
        if 'Total Net Weight (kg)' in pl_result_df.columns:
            pl_result_df['Total Net Weight (kg)'] = pd.to_numeric(pl_result_df['Total Net Weight (kg)'], errors='coerce')
        
        # Fill NaN values with 0 for numerical calculations
        result_df['net weight'] = result_df['net weight'].fillna(0)
        result_df['Total Net Weight (kg)'] = result_df['Total Net Weight (kg)'].fillna(0)
        result_df['Qty'] = result_df['Qty'].fillna(0)
        result_df['Unit Price'] = result_df['Unit Price'].fillna(0)
        
        pl_result_df['Quantity'] = pl_result_df['Quantity'].fillna(0)
        if 'Total Gross Weight (kg)' in pl_result_df.columns:
            pl_result_df['Total Gross Weight (kg)'] = pl_result_df['Total Gross Weight (kg)'].fillna(0)
        if 'Total Net Weight (kg)' in pl_result_df.columns:
            pl_result_df['Total Net Weight (kg)'] = pl_result_df['Total Net Weight (kg)'].fillna(0)
        
        # Calculate total net weight
        net_weight = result_df['net weight']
        total_net_weight = result_df['net weight'].sum()
        print(f"Total net weight calculated: {total_net_weight} kg")
        
        # Safety check for total_net_weight
        if total_net_weight <= 0:
            print("WARNING: Total net weight is zero or negative!")
            # Set a default non-zero value to prevent division by zero later
            total_net_weight = 1
            print(f"Using default weight value: {total_net_weight} kg")
            
    except Exception as e:
        print(f"ERROR in numeric conversion: {e}")
        # Set default values to prevent calculation failures
        total_net_weight = 1
        net_weight = pd.Series([0] * len(result_df))
        print(f"Using default values due to error")

    # Calculate total cost (采购总价) for each row and sum - 保持完整精度
    result_df['采购总价'] = result_df['Unit Price'] * result_df['Qty']
    total_amount = result_df['采购总价'].sum()
    
    # 总价加价就是总价FOB - 保持完整精度
    totalFOB = total_amount * (1 + markup_percentage)

    # 计算总保费 - 保持完整精度
    total_insurance = totalFOB * insurance_coefficient * insurance_rate
    result_df['总保费'] = total_insurance

    result_df['总运费'] = total_freight_amount

    # 修改计算总净重的部分
    try:
        if net_weight_col:
            # 将净重列转换为数值型，保持完整精度
            result_df['net weight'] = pd.to_numeric(result_df['net weight'], errors='coerce')
            result_df['Total Net Weight (kg)'] = pd.to_numeric(result_df['Total Net Weight (kg)'], errors='coerce')
            
            # 修改识别汇总行的逻辑
            total_mask = ~(
                (result_df['NO.'].isna() | (result_df['NO.'] == '')) &
                (result_df['net weight'].notna())
            )
            
            # 计算总净重时只使用非汇总行 - 保持完整精度
            net_weight = result_df[total_mask]['net weight']
            total_net_weight = net_weight.sum()
            
        else:
            print("WARNING: 未找到净重列，使用默认值")
            total_net_weight = 1
            net_weight = pd.Series([0] * len(result_df))
            
        if total_net_weight <= 0:
            print("WARNING: 计算得到的总净重为0或负数！")
            total_net_weight = 1
            print(f"使用默认净重值: {total_net_weight} kg")
            
    except Exception as e:
        print(f"计算总净重时出错: {e}")
        total_net_weight = 1
        net_weight = pd.Series([0] * len(result_df))
        print(f"使用默认值进行计算")

    # 使用正确的总净重继续后续计算 - 保持完整精度
    result_df['每公斤摊的运保费'] = (total_insurance + total_freight_amount) / total_net_weight
    result_df['该项对应的运保费'] = result_df['每公斤摊的运保费'] * result_df['net weight']

    # 总CIF = 总价FOB+总保费+总运费 - 保持完整精度
    total_CIF = totalFOB * (1 + insurance_coefficient * insurance_rate) + total_freight_amount

    # 每公斤净重CIF - 保持完整精度
    cif_per_kg = total_CIF / total_net_weight

    # 每行数据净重*每公斤净重CIF = 该行数据CIF价格 - 保持完整精度
    unit_kg_cif = cif_per_kg * net_weight

    # Calculate FOB price for each item - 保持完整精度
    result_df['采购单价'] = result_df['Unit Price']
    result_df['采购总价'] = result_df['Unit Price'] * result_df['Qty']
    result_df['FOB单价'] = result_df['Unit Price'] * (1 + markup_percentage)
    result_df['FOB总价'] = result_df['FOB单价'] * result_df['Qty']

    # 计算CIF总价和单价 - 保持完整精度
    result_df['CIF总价(FOB总价+运保费)'] = result_df['FOB总价'] + result_df['该项对应的运保费']
    result_df['CIF单价'] = result_df['CIF总价(FOB总价+运保费)'] / result_df['Qty']

    # 计算USD单价 - 保持完整精度
    result_df['单价USD数值'] = result_df['CIF单价'] / exchange_rate

    # 只在最终显示时格式化数值，不影响计算精度
    display_columns = ['采购单价', '采购总价', 'FOB单价', 'FOB总价', '总保费', '总运费', 
                      '每公斤摊的运保费', '该项对应的运保费', 'CIF总价(FOB总价+运保费)', 
                      'CIF单价']
    
    # 这些列在显示时才四舍五入，不影响计算
    for col in display_columns:
        if col in result_df.columns:
            result_df[col] = result_df[col]

    # 确保Unit Price (CIF, USD)保持完整精度
    result_df['单价USD数值'] = result_df['CIF单价'] / exchange_rate

    # Summary statistics
    print(f"\nSummary statistics:")
    print(f"  Total items: {len(result_df)}")
    print(f"  Total net weight: {total_net_weight} kg")  # 不再四舍五入显示
    
    # Calculate unit freight rate (per kg)
    unit_freight_rate = total_freight_amount / total_net_weight if total_net_weight > 0 else 0
    print(f"  Unit freight rate: ¥{unit_freight_rate} per kg")  # 不再四舍五入显示
    print(f"  Markup percentage: {markup_percentage*100}%")  # 不再四舍五入显示
    print(f"  Exchange rate: ¥{exchange_rate} per USD")  # 不再四舍五入显示
    
    # Calculate USD value - 保持原始精度直到最终显示
    result_df['单价USD数值'] = result_df['CIF单价'] * exchange_rate
    
    # Fill in the unit column if it exists
    result_df['单位'] = result_df['Unit'] if 'Unit' in result_df.columns else ""
    
    # Calculate USD unit price
    result_df['Unit Price'] = (result_df['Unit Price'] * exchange_rate).round(8)

    # Calculate Amount as Unit Price multiplied by Quantity
    result_df['Amount'] = (result_df['Unit Price'] * result_df['Qty'] ).round(8)

    # Define output column sets
    cif_output_columns = [
        'NO.', 'Material code', 'DESCRIPTION', 'Model NO.', 'Unit Price', 'Qty', 'Unit', 'Amount',
        'net weight', '采购单价', '采购总价', 'FOB单价', 'FOB总价', '总保费', '总运费', '每公斤摊的运保费',
        '该项对应的运保费', 'CIF总价(FOB总价+运保费)', 'CIF单价', '单价USD数值', '单位',
        'factory', 'project', 'end use'
    ]
    
    cif_output_columns = cif_output_columns + ['G.W (KG)']

    # Define output column sets for export invoice
    exportReimport_output_columns = [
        'S/N', 'Part Number', '名称', 'Model Number', 'Unit Price (CIF, USD)', 'Quantity', 'Unit', 'Total Amount (CIF, USD)', 'Total Net Weight (kg)'
    ]

    # Internal columns needed for calculations (including Trade Type and Shipper)
    internal_columns = cif_output_columns + ['Trade Type', 'Shipper', 'Original_Unit', 'Total Net Weight (kg)']

    # Packing list internal columns
    pl_output_columns.append('project')  # Add project to the output columns but it will be removed before final export

    pl_internal_columns = pl_output_columns + ['Trade Type', 'Shipper', 'factory']
    
    # Ensure all required columns exist
    for col in internal_columns:
        if col not in result_df.columns:
            result_df[col] = None
    
    for col in pl_internal_columns:
        if col not in pl_result_df.columns:
            pl_result_df[col] = None
    
    # Reindex the dataframe to match the required column order for internal processing
    result_df = result_df.reindex(columns=internal_columns)
    pl_result_df = pl_result_df.reindex(columns=pl_internal_columns)

    # Drop rows with no material code or all NaN values
    result_df = result_df.dropna(subset=['Material code'], how='all')
    result_df = result_df.dropna(how='all')
    
    pl_result_df = pl_result_df.dropna(subset=['Part Number'], how='all')
    pl_result_df = pl_result_df.dropna(how='all')
    
    # Apply formatting to numeric columns
    numeric_columns = ['采购单价', '采购总价', 'FOB单价', 'FOB总价', '总保费', '总运费', 
                      '每公斤摊的运保费', '该项对应的运保费', 'CIF总价(FOB总价+运保费)', 
                      'CIF单价', '单价USD数值']
    
    for col in numeric_columns:
        if col in result_df.columns:
            result_df[col] = result_df[col]
    
    # Generate the intermediate CIF invoice file (CIF原始发票)
    cif_invoice = result_df.copy()
    pl_invoice = pl_result_df.copy()
    
    # 确保工厂列在CIF发票中是可见和可用的
    if 'factory' not in cif_invoice.columns or cif_invoice['factory'].isna().all():
        print("WARNING: CIF发票中没有找到有效的工厂列，添加默认工厂")
        cif_invoice['factory'] = '默认工厂'
    
    # 确保项目列在CIF发票中是可见和可用的
    if 'project' not in cif_invoice.columns or cif_invoice['project'].isna().all():
        print("WARNING: CIF发票中没有找到有效的项目列，添加默认项目")
        cif_invoice['project'] = '大华'
    
    # 打印工厂和项目列信息以便调试
    print("\nCIF发票工厂值:")
    if 'factory' in cif_invoice.columns:
        factory_values = cif_invoice['factory'].unique()
        print(f"  工厂唯一值: {factory_values}")
    else:
        print("  未找到工厂列")
    
    print("\nCIF发票项目值:")
    if 'project' in cif_invoice.columns:
        project_values = cif_invoice['project'].unique()
        print(f"  项目唯一值: {project_values}")
    else:
        print("  未找到项目列")
    
    # Remove Trade Type and Shipper columns before saving
    if 'Trade Type' in cif_invoice.columns:
        cif_invoice = cif_invoice.drop(columns=['Trade Type'])
    if 'Shipper' in cif_invoice.columns:
        cif_invoice = cif_invoice.drop(columns=['Shipper'])
    if 'Original_Unit' in cif_invoice.columns:
        cif_invoice = cif_invoice.drop(columns=['Original_Unit'])
        
    if 'Trade Type' in pl_invoice.columns:
        pl_invoice = pl_invoice.drop(columns=['Trade Type'])
    if 'Shipper' in pl_invoice.columns:
        pl_invoice = pl_invoice.drop(columns=['Shipper'])
    if 'Original_Unit' in pl_invoice.columns:
        pl_invoice = pl_invoice.drop(columns=['Original_Unit'])
    if 'factory' in pl_invoice.columns:
        pl_invoice = pl_invoice.drop(columns=['factory'])
    
    # 在保存CIF原始发票之前，确保所有数值列保持完整精度
    numeric_columns = ['采购单价', '采购总价', 'FOB单价', 'FOB总价', '总保费', '总运费', 
                      '每公斤摊的运保费', '该项对应的运保费', 'CIF总价(FOB总价+运保费)', 
                      'CIF单价', '单价USD数值', 'Unit Price', 'Amount']
    
    # 移除所有数值列的格式化，保持原始精度
    for col in numeric_columns:
        if col in cif_invoice.columns:
            # 确保数值类型，但不进行任何四舍五入
            cif_invoice[col] = pd.to_numeric(cif_invoice[col], errors='coerce')
    
    cif_file_path = os.path.join(output_dir, 'cif_original_invoice.xlsx')
    pl_file_path = os.path.join(output_dir, 'pl_original_invoice.xlsx')
    
    # 保存CIF发票时不进行任何格式化或四舍五入
    with pd.ExcelWriter(cif_file_path, engine='openpyxl') as writer:
        cif_invoice.to_excel(writer, index=False)
        
        # 获取工作表
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # 设置数值列的格式以显示完整精度
        for col_idx, col_name in enumerate(cif_invoice.columns, 1):
            if col_name in numeric_columns:
                # 使用自定义数字格式来显示所有有效数字
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(cif_invoice) + 2):  # 从第2行开始（跳过表头）
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = '0.############'  # 使用足够多的#来显示所有有效数字
    
    # 保存打包清单
    safe_save_to_excel(pl_invoice, pl_file_path)
    
    # 提取一般贸易的物料
    general_trade_df = result_df[result_df['Trade Type'] == '一般贸易'].copy()
    pl_df = pl_result_df[pl_result_df['Trade Type'] == '一般贸易'].copy()

    print(f"\nGeneral trade count in result_df: {len(general_trade_df)}")
    print(f"General trade count in pl_result_df: {len(pl_df)}")
    
    print("\nColumns in pl_result_df:")
    print(pl_result_df.columns.tolist())
    
    print("\nColumns in pl_df:")
    print(pl_df.columns.tolist() if not pl_df.empty else "pl_df is empty")
    
    # 只有在存在一般贸易物料时才生成出口发票文件
    if not general_trade_df.empty:
        # Generate the export invoice with two sheets - packing list and commercial invoice
        # First, create a copy for packing list (Sheet1)
        packing_list = general_trade_df.copy()
        
        # Remove Trade Type, Shipper, and project columns before saving to Excel
        if 'Trade Type' in packing_list.columns:
            packing_list = packing_list.drop(columns=['Trade Type'])
        if 'Shipper' in packing_list.columns:
            packing_list = packing_list.drop(columns=['Shipper'])
        if 'project' in packing_list.columns:
            packing_list = packing_list.drop(columns=['project'])
        
        # Use original Chinese units for export invoice
        export_invoice = general_trade_df[['NO.', 'Material code', 'DESCRIPTION', 'Model NO.', '单价USD数值', 'Qty', 'Unit', 'Amount', 'Total Net Weight (kg)']].copy()
        
        # 将人民币单价转换为美元单价（除以汇率）
        export_invoice['Unit Price (CIF, USD)'] = export_invoice['单价USD数值'].round(4)
        
        # 重命名列（但保持 Unit Price (CIF, USD) 不变，因为我们已经直接设置了这个列）
        export_invoice.rename(columns={
            'NO.': 'S/N',
            'Material code': 'Part Number',
            'DESCRIPTION': '名称',
            'Model NO.': 'Model Number',
            'Qty': 'Quantity',
            'Amount': 'Total Amount (CIF, USD)'
        }, inplace=True)
        
        # 重新计算美元总金额
        export_invoice['Total Amount (CIF, USD)'] = (export_invoice['Unit Price (CIF, USD)'] * export_invoice['Quantity']).round(4)
        
        # Keep original Chinese units from the source
        if 'Original_Unit' in general_trade_df.columns:
            export_invoice['Unit'] = general_trade_df['Original_Unit']
        
        # Group by Material code and other fields, keeping original units
        export_grouped = export_invoice.groupby(['Part Number', 'Unit Price (CIF, USD)'], as_index=False).agg({
            'Quantity': 'sum',
            'S/N': 'first',
            'Unit': 'first',  # This will keep the original Chinese unit
            'Model Number': 'first',
            '名称': 'first',
            'Total Net Weight (kg)': 'sum',  # Sum the total net weight for grouped items
        })
        
        # Calculate Amount after grouping
        export_grouped['Total Amount (CIF, USD)'] = (export_grouped['Unit Price (CIF, USD)'] * export_grouped['Quantity'])
        
        # Ensure all required columns exist and in correct order
        for col in exportReimport_output_columns:
            if col not in export_grouped.columns:
                export_grouped[col] = None
        
        # Reindex to match the required column order
        export_grouped = export_grouped.reindex(columns=exportReimport_output_columns)
        
        # Sort by S/N to maintain original ordering
        export_grouped = export_grouped.sort_values('S/N')
        
        # Reset the index to generate sequential numbers
        export_grouped = export_grouped.reset_index(drop=True)
        export_grouped['S/N'] = export_grouped.index + 1

        # Save both sheets to the same Excel file
        export_file_path = os.path.join(output_dir, 'export_invoice.xlsx')
        
        # Try to delete existing file to avoid permission issues
        try:
            if os.path.exists(export_file_path):
                os.remove(export_file_path)
                print(f"Removed existing file: {export_file_path}")
                time.sleep(1)  # Give the OS time to fully release the file
        except Exception as e:
            print(f"Warning: Could not remove existing file: {e}")
        
        # Create a new Excel writer
        with pd.ExcelWriter(export_file_path, engine='openpyxl') as writer:
            # Packing List 工作表处理
            if not pl_df.empty:
                packing_df = pl_df.copy()
                
                # 确保所有需要的列都存在
                for col in pl_output_columns:
                    if col not in packing_df.columns:
                        print(f"Warning: Column '{col}' not found in packing list data, adding empty column")
                        packing_df[col] = None
                
                # 移除 project 列
                if 'project' in packing_df.columns:
                    packing_df = packing_df.drop(columns=['project'])
                
                # 确保正确的输出列顺序（不包含 project）
                output_columns = [col for col in pl_output_columns if col != 'project']
                packing_df = packing_df[output_columns]
                
                # 添加汇总行（只对数字列计算总和）
                summary_cols = ['Quantity', 'Total Gross Weight (kg)', 'Total Net Weight (kg)', 'Total Carton Quantity', 'Total Volume (CBM)']
                summary_packing = {'名称': 'Total'}
                for col in summary_cols:
                    if col in packing_df.columns:
                        # Calculate sum without modifying the original column in place
                        # Coerce to numeric, fill NA with 0 JUST for the sum calculation
                        summary_packing[col] = pd.to_numeric(packing_df[col], errors='coerce').fillna(0).sum()

                summary_row = pd.DataFrame([{col: (summary_packing.get(col, None) if col in summary_cols else None) for col in packing_df.columns}])
                summary_row['名称'] = 'Total'
                packing_df = pd.concat([packing_df, summary_row], ignore_index=True)
                
                # Debug print columns
                print("\nPacking List columns before saving:")
                print(packing_df.columns.tolist())
                
                # 添加PL页脚信息
                # 获取包裹数量
                total_packages = int(summary_packing.get('Total Carton Quantity', 0))
                # 获取总净重
                total_net_weight = summary_packing.get('Total Net Weight (kg)', 0)
                # 获取总毛重
                total_gross_weight = summary_packing.get('Total Gross Weight (kg)', 0)
                # 获取总体积
                total_volume = summary_packing.get('Total Volume (CBM)', 0)
                
                # 创建页脚行
                footer_rows = [
                    {'S/N': f'PACKED IN {total_packages} PACKAGES ONLY.'},
                    {'S/N': f'NET WEIGHT: {total_net_weight:.2f} KGS'},
                    {'S/N': f'GROSS WEIGHT: {total_gross_weight:.2f} KGS'},
                    {'S/N': f'TOTAL MEASUREMENT:{total_volume:.2f} CBM'},
                    {'S/N': 'COUNTRY OF ORIGIN: CHINA'}
                ]
                
                # 为每行添加空白列，确保列数匹配
                for row in footer_rows:
                    for col in packing_df.columns:
                        if col not in row:
                            row[col] = None
                
                # 将页脚行添加到数据框
                footer_df = pd.DataFrame(footer_rows)
                packing_df = pd.concat([packing_df, footer_df], ignore_index=True)
                
                # 保存到Excel
                packing_df.to_excel(writer, sheet_name='PL', index=False)
                
                # Apply cell merging for packing list
                merge_packing_list_cells(export_file_path)
                
                # Apply footer styling for packing list
                apply_pl_footer_styling(export_file_path)
            else:
                # 如果没有pl_df数据，创建一个空的packing list with correct columns (不包含 project)
                empty_pl_df = pd.DataFrame(columns=[col for col in pl_output_columns if col != 'project'])
                empty_pl_df.to_excel(writer, sheet_name='PL', index=False)

            # Commercial Invoice 工作表处理
            commercial_df = export_grouped.copy()
            # 添加汇总行
            summary_commercial = {'名称': 'Total', 'Part Number': ''}
            for col in ['Quantity', 'Total Amount (CIF, USD)', 'Total Net Weight (kg)']:
                if col in commercial_df.columns:
                    summary_commercial[col] = pd.to_numeric(commercial_df[col], errors='coerce').fillna(0).sum()
            
            # Create new row with just the sums for Qty and Amount
            summary_row = pd.DataFrame([summary_commercial])
            
            # Proper column order for the summary row
            for col in exportReimport_output_columns:
                if col not in summary_row.columns:
                    summary_row[col] = None
            
            summary_row = summary_row[exportReimport_output_columns]
            
            # Get the total amount from the summary row
            total_amount = summary_commercial.get('Total Amount (CIF, USD)', 0)
            total_amount_words = num_to_words(total_amount)
            
            # Format following the screenshot: "SAY USD [AMOUNT IN WORDS] ONLY."
            # Create an empty row with all fields blank, but spanning all columns
            empty_row_data = {col: "" for col in exportReimport_output_columns}
            empty_row = pd.DataFrame([empty_row_data])
            
            # Create the words row with the exact format from screenshot
            words_row_data = {col: "" for col in exportReimport_output_columns}
            words_row_data['S/N'] = f"Amount in Words: SAY USD {total_amount_words} ONLY."
            words_row = pd.DataFrame([words_row_data])
            
            # Add both rows to the DataFrame (summary row, empty row, words row)
            commercial_df = pd.concat([commercial_df, summary_row, empty_row, words_row], ignore_index=True)
            
            # 使用正确的发票号码格式作为工作表名
            invoice_sheet_name = generate_invoice_sheet_name()
            print(f"Using invoice sheet name: {invoice_sheet_name}")
            commercial_df.to_excel(writer, sheet_name=invoice_sheet_name, index=False)

            # 确保至少一个工作表可见
            workbook = writer.book
            for sheet in workbook.worksheets:
                sheet.sheet_state = "visible"  # 显式设置所有工作表可见
            
            # 设置默认打开发票工作表
            worksheet = workbook[invoice_sheet_name]
            workbook.active = workbook.index(worksheet)
        
        # Apply styling to both sheets
        try:
            # Load the workbook
            wb = load_workbook(export_file_path)
            
            # Style each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Define styles
                header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Apply styling to headers
                for col_idx, col in enumerate(ws[1], 1):
                    col.font = header_font
                    col.fill = header_fill
                    col.alignment = header_alignment
                    
                    # Set column width - customize widths for specific columns
                    col_name = ws.cell(row=1, column=col_idx).value
                    if col_name == 'Part Number':
                        ws.column_dimensions[get_column_letter(col_idx)].width = 35  # Wider for Part Number
                    elif col_name == 'Unit Price (CIF, USD)':
                        ws.column_dimensions[get_column_letter(col_idx)].width = 25  # Wider for Unit Price
                    elif col_name == 'Total Amount (CIF, USD)':
                        ws.column_dimensions[get_column_letter(col_idx)].width = 25  # Wider for Amount
                    elif col_name == '名称':
                        ws.column_dimensions[get_column_letter(col_idx)].width = 30  # Wider for Description
                    elif col_name == 'Model Number':
                        ws.column_dimensions[get_column_letter(col_idx)].width = 20  # Wider for Model Number
                    elif col_name == 'Total Net Weight (kg)':
                        ws.column_dimensions[get_column_letter(col_idx)].width = 20  # Wider for Net Weight
                    else:
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15  # Default width
                
                # Apply borders to all cells
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                
                # Freeze the header row
                ws.freeze_panes = 'A2'
            
            # Apply number formatting to specific columns
            unit_price_col = None
            amount_col = None
            
            # Find the Unit Price and Amount column indices
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Unit Price (CIF, USD)':
                    unit_price_col = col_idx
                elif cell.value == 'Total Amount (CIF, USD)':
                    amount_col = col_idx
            
            # Apply formatting to the entire column
            if unit_price_col:
                for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                    cell = ws.cell(row=row, column=unit_price_col)
                    cell.number_format = '#,##0.0000'
            
            if amount_col:
                for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                    cell = ws.cell(row=row, column=amount_col)
                    cell.number_format = '#,##0.00'
            
            # Find the "SAY USD" row and merge cells to span across all columns
            if sheet_name != 'PL':  # Only for invoice sheets, not packing list
                for row_idx in range(1, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value == "SAY USD":
                        # Get the total amount words from the Part Number column
                        amount_words = ws.cell(row=row_idx, column=2).value
                        only_text = ws.cell(row=row_idx, column=3).value if ws.cell(row=row_idx, column=3).value else "ONLY."
                        # Create the full text in the proper format
                        full_text = f"SAY USD {amount_words} {only_text}"
                        # Set this text in the first cell
                        ws.cell(row=row_idx, column=1).value = full_text
                        # Merge cells across all columns
                        last_col = len(exportReimport_output_columns)
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
                        # Align the text left
                        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
                        break
            
            # Find the "Amount in Words:" row and merge cells to span across all columns
            if sheet_name != 'PL':  # Only for invoice sheets, not packing list
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=1).value
                    if cell_value and "Amount in Words:" in str(cell_value):
                        # Merge all columns in this row
                        last_col = len(exportReimport_output_columns)
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
                        # Align the text left and make it bold
                        cell = ws.cell(row=row_idx, column=1)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.font = Font(bold=True)

                        # Add company information rows after Amount in Words
                        company_info = [
                            f"Country Of Origin: ",
                            f"Payment Term: ",
                            f"Delivery Term: ",
                            f"Company Name: {pc}",
                            f"Account number: {ba}",
                            f"Bank Name: {bn}",
                            f"Bank Address: {badd}",
                            f"SWIFT No.: {swn}"
                        ]

                        # Start row for info block
                        info_start_row = row_idx + 1

                        # Add company info rows
                        for i, info in enumerate(company_info):
                            # Merge cells for each info row
                            ws.merge_cells(start_row=info_start_row + i, start_column=1, end_row=info_start_row + i, end_column=last_col)
                            # Add the info text
                            cell = ws.cell(row=info_start_row + i, column=1)
                            cell.value = info
                            cell.alignment = Alignment(horizontal='left', vertical='center')

                        # Add signature and date fields
                        signature_row = info_start_row + len(company_info)
                        ws.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=last_col)
                        signature_cell = ws.cell(row=signature_row, column=1)
                        signature_cell.value = ""
                        signature_cell.alignment = Alignment(horizontal='right', vertical='center')

                        # Add empty row after signature
                        empty_row = signature_row + 1
                        ws.merge_cells(start_row=empty_row, start_column=1, end_row=empty_row, end_column=last_col)
                        ws.cell(row=empty_row, column=1).value = ""

                        # Do not repeat the company info block again

                    # Make all text in the sheet bold
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value:  # Only apply bold to cells with content
                                current_font = cell.font
                                cell.font = Font(
                                    name=current_font.name,
                                    size=current_font.size,
                                    bold=True,
                                    italic=current_font.italic,
                                    color=current_font.color
                                )

                    # Apply selective bold formatting
                    apply_selective_bold(ws)
            
            # Save the styled workbook
            wb.save(export_file_path)
            
            # Apply cell merging for packing list
            merge_packing_list_cells(export_file_path)
            
            # Apply footer styling for packing list
            apply_pl_footer_styling(export_file_path)
            
            print(f"Successfully saved and styled export file with multiple sheets: {export_file_path}")
            
            # After saving the export_invoice.xlsx, now merge it with h.xlsx and f.xlsx
            print("Merging files: h.xlsx, export_invoice.xlsx, f.xlsx")
            
            # Temporarily save export_invoice.xlsx to a backup file
            temp_export_file = os.path.join(output_dir, 'temp_export_invoice.xlsx')
            try:
                import shutil
                shutil.copy(export_file_path, temp_export_file)
                
                # Prepare file paths for merging
                h_file = 'h.xlsx'
                f_file = 'f.xlsx'
                
                # For Packing List sheet, we need different files
                pl_h_file = 'pl_h.xlsx'  # First file for Packing List sheet
                pl_f_file = 'pl_f.xlsx'  # Last file for Packing List sheet
                
                # Function to find a file in multiple locations
                def find_file(filename):
                    # Check in current directory
                    if os.path.exists(filename):
                        return filename
                    
                    # Check in output_dir
                    file_in_output = os.path.join(output_dir, filename)
                    if os.path.exists(file_in_output):
                        return file_in_output
                    
                    # Check in the script's directory
                    script_dir = os.path.dirname(os.path.abspath(__file__))
                    file_in_script_dir = os.path.join(script_dir, filename)
                    if os.path.exists(file_in_script_dir):
                        return file_in_script_dir
                    
                    return None  # File not found
                
                # Find the files
                h_file_path = find_file(h_file)
                f_file_path = find_file(f_file)
                pl_h_file_path = find_file(pl_h_file)
                pl_f_file_path = find_file(pl_f_file)
                
                if h_file_path and f_file_path:
                    print(f"Found files for merging Commercial Invoice:")
                    print(f"  {h_file}: {h_file_path}")
                    print(f"  {f_file}: {f_file_path}")
                    
                    if pl_h_file_path and pl_f_file_path:
                        print(f"Found files for merging Packing List:")
                        print(f"  {pl_h_file}: {pl_h_file_path}")
                        print(f"  {pl_f_file}: {pl_f_file_path}")
                    else:
                        print("Warning: Packing List merge files not found, will only merge Commercial Invoice")
                    
                    # Import functions from merge.py
                    import importlib.util
                    import sys
                    
                    # Get the path to merge.py (checking multiple locations)
                    merge_py_path = find_file('merge.py')
                    
                    if merge_py_path:
                        print(f"Found merge.py at: {merge_py_path}")
                        
                        # Before merging, let's rename our sheet in the temporary file to match the invoice sheet name
                        try:
                            # Make a backup of the temporary file
                            book = load_workbook(temp_export_file)
                            
                            # Check if 'Commercial Invoice' sheet exists and rename it to invoice_sheet_name
                            if 'Commercial Invoice' in book.sheetnames:
                                print(f"Renaming 'Commercial Invoice' sheet to '{invoice_sheet_name}' before merging")
                                ci_sheet = book['Commercial Invoice']
                                ci_sheet.title = invoice_sheet_name
                                book.save(temp_export_file)
                            
                        except Exception as e:
                            print(f"Warning: Could not rename sheet in temporary file: {e}")
                        
                        # Now continue with the merge operation
                        # Call merge.py with the files in specific order using subprocess
                        import subprocess
                        
                        # Prepare command with or without Packing List files
                        if pl_h_file_path and pl_f_file_path:
                            merge_cmd = [
                                sys.executable, 
                                merge_py_path, 
                                h_file_path, 
                                temp_export_file, 
                                f_file_path, 
                                export_file_path,
                                pl_h_file_path,
                                pl_f_file_path
                            ]
                        else:
                            merge_cmd = [
                                sys.executable, 
                                merge_py_path, 
                                h_file_path, 
                                temp_export_file, 
                                f_file_path, 
                                export_file_path
                            ]
                            
                        print(f"Running merge command: {' '.join(merge_cmd)}")
                        
                        try:
                            # Use subprocess.run with stdout and stderr captured to diagnose issues
                            result = subprocess.run(
                                merge_cmd, 
                                check=True,
                                capture_output=True,
                                text=True
                            )
                            
                            # Print stdout and stderr for debugging
                            if result.stdout:
                                print("Merge output:")
                                print(result.stdout)
                            
                            if result.stderr:
                                print("Merge errors:")
                                print(result.stderr)
                            
                            # If successful, verify the file exists
                            if os.path.exists(export_file_path):
                                print(f"Successfully merged files into: {export_file_path}")
                            else:
                                print(f"Error: Merged file not created at {export_file_path}")
                        except subprocess.CalledProcessError as e:
                            print(f"Error running merge script: {e}")
                            if hasattr(e, 'stderr') and e.stderr:
                                print(f"Error details: {e.stderr}")
                        except Exception as e:
                            print(f"Error during file merging: {e}")
                            # Restore original export_invoice.xlsx if merging failed
                            if os.path.exists(temp_export_file):
                                shutil.copy(temp_export_file, export_file_path)
                                print("Restored original export_invoice.xlsx")
                    else:
                        print(f"Error: merge.py not found")
                else:
                    missing_files = []
                    if not h_file_path:
                        missing_files.append(h_file)
                    if not f_file_path:
                        missing_files.append(f_file)
                    if missing_files:
                        print(f"Warning: Could not merge files. Missing files: {', '.join(missing_files)}")
                    else:
                        print("Unexpected error: All files found but merging could not proceed")
            except Exception as e:
                print(f"Error during file merging: {e}")
            finally:
                # Clean up temporary file
                if os.path.exists(temp_export_file):
                    try:
                        os.remove(temp_export_file)
                    except:
                        pass
        except Exception as e:
            print(f"Warning: File saved but could not apply styling: {e}")
    else:
        print("没有一般贸易的物料，不生成出口发票文件")
    
    # After creating result_df and before generating any output files
    # Split the data by project and factory
    split_dfs, project_categories = split_by_project_and_factory(result_df)
    
    # Generate a single invoice file with multiple sheets for all splits
    reimport_invoice_path = os.path.join(output_dir, 'reimport_invoice.xlsx')
    
    # Delete existing reimport_invoice.xlsx file if it exists
    if os.path.exists(reimport_invoice_path):
        try:
            os.remove(reimport_invoice_path)
            print(f"Removed existing file: {reimport_invoice_path}")
            time.sleep(1)  # Give the OS time to fully release the file
        except Exception as e:
            print(f"Warning: Could not remove existing file: {e}")
    
    # Create a new Excel writer for the reimport file
    with pd.ExcelWriter(reimport_invoice_path, engine='openpyxl') as writer:
        # First, add the complete Packing List sheet
        complete_pl_df = pl_result_df.copy()
        # Remove internal columns before saving
        save_columns = [col for col in pl_output_columns if col != 'project']  # Remove project from output
        complete_pl_df = complete_pl_df[save_columns]
        
        # Add summary row to packing list
        summary_cols = ['Quantity', 'Total Gross Weight (kg)', 'Total Net Weight (kg)', 'Total Carton Quantity', 'Total Volume (CBM)']
        summary_packing = {'名称': 'Total'}
        for col in summary_cols:
            if col in complete_pl_df.columns:
                # Calculate sum without modifying the original column in place
                # Coerce to numeric, fill NA with 0 JUST for the sum calculation
                summary_packing[col] = pd.to_numeric(complete_pl_df[col], errors='coerce').fillna(0).sum()
        
        summary_row = pd.DataFrame([summary_packing])
        complete_pl_df = pd.concat([complete_pl_df, summary_row], ignore_index=True)
        
        # 添加PL页脚信息
        # 获取包裹数量
        total_packages = int(summary_packing.get('Total Carton Quantity', 0))
        # 获取总净重
        total_net_weight = summary_packing.get('Total Net Weight (kg)', 0)
        # 获取总毛重
        total_gross_weight = summary_packing.get('Total Gross Weight (kg)', 0)
        # 获取总体积
        total_volume = summary_packing.get('Total Volume (CBM)', 0)
        
        # 创建页脚行
        footer_rows = [
            {'S/N': f'PACKED IN {total_packages} PACKAGES ONLY.'},
            {'S/N': f'NET WEIGHT: {total_net_weight:.2f} KGS'},
            {'S/N': f'GROSS WEIGHT: {total_gross_weight:.2f} KGS'},
            {'S/N': f'TOTAL MEASUREMENT:{total_volume:.2f} CBM'},
            {'S/N': 'COUNTRY OF ORIGIN: CHINA'}
        ]
        
        # 为每行添加空白列，确保列数匹配
        for row in footer_rows:
            for col in complete_pl_df.columns:
                if col not in row:
                    row[col] = None
        
        # 将页脚行添加到数据框
        footer_df = pd.DataFrame(footer_rows)
        complete_pl_df = pd.concat([complete_pl_df, footer_df], ignore_index=True)
        
        # Save packing list sheet
        complete_pl_df.to_excel(writer, sheet_name='PL', index=False)
        
        # Process each split for Commercial Invoice sheets only
        # Generate base invoice number and increment for each sheet
        base_invoice_name = generate_invoice_sheet_name(prefix="RIMP")
        # Extract the numeric part for incrementing 
        invoice_prefix = base_invoice_name[:-4]  # Everything except last 4 digits
        invoice_number = int(base_invoice_name[-4:])  # Last 4 digits as integer
        
        print(f"Base invoice name: {base_invoice_name}")
        print(f"Prefix: {invoice_prefix}, Starting number: {invoice_number}")
        
        # Sort keys to ensure consistent ordering
        sorted_keys = sorted(split_dfs.keys())
        
        # Track sheet names being created
        created_sheet_names = []
        
        for key in sorted_keys:
            project, factory = key
            df = split_dfs[key]
            
            if not df.empty:
                # Create sequential invoice sheet name
                ci_sheet_name = f"{invoice_prefix}{invoice_number:04d}"
                created_sheet_names.append(ci_sheet_name)
                print(f"Using sheet name '{ci_sheet_name}' for project '{project}', factory '{factory}'")
                
                # Increment for next sheet
                invoice_number += 1
                
                # 确保文件名中的工厂和项目值是有效的字符串
                project_safe = str(project).strip().replace(' ', '_')
                factory_safe = str(factory).strip().replace(' ', '_')
                
                # 创建独立的进口发票文件
                reimport_file_name = f'reimport_{project_safe}_{factory_safe}.xlsx'
                reimport_file_path = os.path.join(output_dir, reimport_file_name)
                
                # Create a copy for the invoice
                invoice_df = df[['NO.', 'Material code', 'DESCRIPTION', 'Model NO.', 'CIF单价', 'Qty', 'Unit', 'CIF总价(FOB总价+运保费)', 'Total Net Weight (kg)']].copy()
                # 替换"名称"为"Commodity Description (Customs)"，并赋值
                if customs_desc_col is not None and customs_desc_col in packing_list_df.columns:
                    invoice_df['Commodity Description (Customs)'] = packing_list_df.loc[df.index, customs_desc_col].values
                else:
                    invoice_df['Commodity Description (Customs)'] = invoice_df['DESCRIPTION']
                # 删除"DESCRIPTION"列
                invoice_df.drop(columns=['DESCRIPTION'], inplace=True)
                # 调整列顺序，确保Commodity Description (Customs)在Part Number后面
                reimport_columns = [
                    'S/N', 'Part Number', 'Commodity Description (Customs)', 'Model Number', 'Unit Price (CIF, USD)', 'Quantity', 'Unit', 'Total Amount (CIF, USD)', 'Total Net Weight (kg)'
                ]
                invoice_df = invoice_df.rename(columns={
                    'NO.': 'S/N',
                    'Material code': 'Part Number',
                    'Model NO.': 'Model Number',
                    'CIF单价': 'Unit Price (CIF, USD)',
                    'Qty': 'Quantity',
                    'CIF总价(FOB总价+运保费)': 'Total Amount (CIF, USD)',
                    'Unit': 'Unit'
                })
                invoice_df = invoice_df[reimport_columns]
                # Add summary row to invoice
                summary_invoice = {col: '' for col in reimport_columns}
                summary_invoice['Commodity Description (Customs)'] = 'Total'
                summary_invoice['Part Number'] = ''
                for col in ['Quantity', 'Total Amount (CIF, USD)', 'Total Net Weight (kg)']:
                    if col in invoice_df.columns:
                        summary_invoice[col] = pd.to_numeric(invoice_df[col], errors='coerce').fillna(0).sum()
                summary_row = pd.DataFrame([summary_invoice])[reimport_columns]
                # Create empty row and words row
                empty_row = pd.DataFrame([{col: '' for col in reimport_columns}])
                words_row = pd.DataFrame([{col: '' for col in reimport_columns}])
                words_row['S/N'] = 'Amount in Words:'
                words_row['Part Number'] = f"SAY USD {total_amount_words} ONLY."
                # Add all rows to the DataFrame
                invoice_df = pd.concat([invoice_df, summary_row, empty_row, words_row], ignore_index=True)[reimport_columns]
                
                # Save as individual reimport file
                print(f"Saving individual reimport file for {project}_{factory}: {reimport_file_path}")
                safe_save_to_excel(invoice_df, reimport_file_path)
                
                # Save to combined workbook
                invoice_df.to_excel(writer, sheet_name=ci_sheet_name, index=False)
                
                print(f"Added Commercial Invoice sheet for project {project}, factory {factory}")
        
        print(f"Created reimport invoice sheets: {created_sheet_names}")
    
    # Verify the sheet names in the saved file
    try:
        verification_xls = pd.ExcelFile(reimport_invoice_path)
        print(f"VERIFICATION - Sheets in saved reimport_invoice.xlsx: {verification_xls.sheet_names}")
    except Exception as e:
        print(f"Error verifying sheet names: {e}")
    
    # Apply styling to the reimport file
    try:
        # Load the workbook
        wb = load_workbook(reimport_invoice_path)
        
        # Style each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Define styles
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Apply styling to headers
            for col_idx, col in enumerate(ws[1], 1):
                col.font = header_font
                col.fill = header_fill
                col.alignment = header_alignment
                
                # Set column width - customize widths for specific columns
                col_name = ws.cell(row=1, column=col_idx).value
                if col_name == 'Part Number':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 35  # Wider for Part Number
                elif col_name == 'Unit Price (CIF, USD)':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 25  # Wider for Unit Price
                elif col_name == 'Total Amount (CIF, USD)':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 25  # Wider for Amount
                elif col_name == '名称':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 30  # Wider for Description
                elif col_name == 'Model Number':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20  # Wider for Model Number
                elif col_name == 'Total Net Weight (kg)':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 20  # Wider for Net Weight
                else:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15  # Default width
            
            # Apply borders to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
            # Freeze the header row
            ws.freeze_panes = 'A2'
            
            # Apply selective bold formatting
            apply_selective_bold(ws)
            
            # Apply number formatting to specific columns if this is a Commercial Invoice sheet
            if sheet_name != 'PL':
                for col_idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Unit Price (CIF, USD)':
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.number_format = '#,##0.0000'
                    elif cell.value == 'Total Amount (CIF, USD)':
                        for row in range(2, ws.max_row + 1):
                            cell = ws.cell(row=row, column=col_idx)
                            cell.number_format = '#,##0.00'
                
                # Find the "SAY USD" row and merge cells to span all columns
                for row_idx in range(1, ws.max_row + 1):
                    if ws.cell(row=row_idx, column=1).value == "SAY USD":
                        # Get the total amount words from the Part Number column
                        amount_words = ws.cell(row=row_idx, column=2).value
                        only_text = ws.cell(row=row_idx, column=3).value if ws.cell(row=row_idx, column=3).value else "ONLY."
                        # Create the full text in the proper format
                        full_text = f"SAY USD {amount_words} {only_text}"
                        # Set this text in the first cell
                        ws.cell(row=row_idx, column=1).value = full_text
                        # Merge cells across all columns
                        last_col = len(exportReimport_output_columns)
                        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
                        # Align the text left
                        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='left', vertical='center')
                        break
        
        # Save the styled workbook
        wb.save(reimport_invoice_path)
        merge_packing_list_cells(reimport_invoice_path)
        
        # Apply footer styling for packing list
        apply_pl_footer_styling(reimport_invoice_path)
        
        # Save the final results
        print(f"Successfully generated all files in {output_dir}:")
        print(f"1. {os.path.basename(export_file_path)}")
        
    except Exception as e:
        print(f"Warning: Could not apply styling to reimport invoice file: {e}")

    # 添加验证步骤
    if 'G.W (KG)' in pl_result_df.columns:
        # 检查是否有异常的重量值
        gw_values = pd.to_numeric(pl_result_df['G.W (KG)'], errors='coerce')
        if not gw_values.empty and gw_values.max() < gw_values.sum() * 0.5:  # 如果最大值远小于总和，可能使用了单件重量
            print("WARNING: G.W (KG) values seem too small, might be using unit weights instead of total weights")
            print(f"Max weight: {gw_values.max()}, Total weight: {gw_values.sum()}")

    return result_df

# Run the process
if __name__ == "__main__":
    # Set up command line argument parsing
    parser = argparse.ArgumentParser(description='处理装运清单并生成出口和复进口发票')
    
    parser.add_argument('--packing-list', type=str, default='testfiles/original_packing_list.xlsx',
                      help='原始装箱单文件路径 (默认: testfiles/original_packing_list.xlsx)')
    
    parser.add_argument('--policy', type=str, default='testfiles/policy.xlsx',
                      help='政策文件路径 (默认: testfiles/policy.xlsx)')
    
    parser.add_argument('--output-dir', type=str, default='outputs',
                      help='输出目录 (默认: outputs)')
    
    parser.add_argument('--debug', action='store_true', 
                      help='启用调试模式，显示详细错误信息')
                      
    args = parser.parse_args()
    
    # Create output directory if it doesn't exist
    if not os.path.exists(args.output_dir):
        try:
            os.makedirs(args.output_dir)
            print(f"Created output directory: {args.output_dir}")
        except Exception as e:
            print(f"Error creating output directory: {e}")
            raise
    
    # Get file paths from arguments
    packing_list_file = args.packing_list
    policy_file = args.policy
    
    try:
        # Verify input files exist
        if not os.path.exists(packing_list_file):
            raise FileNotFoundError(f"原始装箱单文件不存在: {packing_list_file}")
        
        if not os.path.exists(policy_file):
            raise FileNotFoundError(f"政策文件不存在: {policy_file}")
        
        # Check file formats
        if not packing_list_file.lower().endswith('.xlsx'):
            print(f"警告: 装箱单文件 '{packing_list_file}' 可能不是Excel格式")
        
        if not policy_file.lower().endswith('.xlsx'):
            print(f"警告: 政策文件 '{policy_file}' 可能不是Excel格式")
        
        print(f"开始处理文件:")
        print(f"- 装箱单: {packing_list_file}")
        print(f"- 政策文件: {policy_file}")
        print(f"- 输出目录: {args.output_dir}")
        
        result = process_shipping_list(packing_list_file, policy_file, args.output_dir)
        print(f"处理完成！输出文件已保存到 '{args.output_dir}' 目录。")
    except FileNotFoundError as e:
        print(f"错误: {e}")
    except Exception as e:
        print(f"处理文件时出错: {e}")
        
        # 打印详细错误信息
        if args.debug:
            import traceback
            traceback.print_exc()
        else:
            import traceback
            print(f"错误位置: {traceback.format_exc().splitlines()[-2]}")
            print("使用 --debug 参数可以查看详细错误信息")
        
        # 打印列名进行调试
        try:
            # Test reading with different skiprows
            for skip in [0, 1, 2, 3]:
                try:
                    print(f"\n尝试跳过 {skip} 行读取政策文件:")
                    policy_df = pd.read_excel(policy_file, skiprows=skip, nrows=5)
                    print(f"前5行: {policy_df.head().values.tolist()}")
                    print(f"列名: {list(policy_df.columns)}")
                except Exception as skip_err:
                    print(f"  - 跳过 {skip} 行时出错: {skip_err}")
        except Exception as read_err:
            print(f"读取政策文件时出错: {read_err}")
            
        try:
            print("\n装箱单文件结构:")
            packing_list_peek = pd.read_excel(packing_list_file, nrows=5, header=None)
            for i, row in enumerate(packing_list_peek.values.tolist()):
                print(f"第 {i+1} 行: {row[:5]}...")
            
            # Try reading with different header configurations
            print("\n尝试不同的表头配置读取装箱单:")
            
            # Standard read
            try:
                print("\n标准读取:")
                packing_list_df = pd.read_excel(packing_list_file)
                print(f"列名: {list(packing_list_df.columns)[:5]}...")
            except Exception as e:
                print(f"标准读取出错: {e}")
                
            # With header=[1,2]
            try:
                print("\n多级表头读取 [第2-3行]:")
                packing_list_df = pd.read_excel(packing_list_file, header=[1,2])
                print(f"列名: {list(packing_list_df.columns)[:5]}...")
            except Exception as e:
                print(f"多级表头读取出错: {e}")
                
            # Skip first row, use header=[0,1]
            try:
                print("\n跳过首行，多级表头读取:")
                packing_list_df = pd.read_excel(packing_list_file, header=[0,1], skiprows=[0])
                print(f"列名: {list(packing_list_df.columns)[:5]}...")
            except Exception as e:
                print(f"跳过首行多级表头读取出错: {e}")
                
        except Exception as peek_err:
            print(f"检查装箱单文件结构时出错: {peek_err}")